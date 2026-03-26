"""
pointage_pivot.py
-----------------
Reads one or more MEDIDIS "ETAT DES HEURES TRAVAILLEES" Excel files and
produces a styled pivot table (one sheet per input file + a merged sheet
if multiple files) showing, per employee:
  - Monthly hours worked  (first entry → last exit, ignoring intermediate punches)
  - Monthly absence hours (days with no valid punches × standard daily hours)
  - Annual totals

Usage:
    python pointage_pivot.py file1.xlsx [file2.xlsx ...] [-o output.xlsx]

Defaults:
    Input  : All .xlsx files in the input/ directory
    Output : One .xlsx file per input file, named <input>_pivot.xlsx
"""

import os
import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MONTHS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]

# Standard working hours per day (adjust if needed)
STANDARD_DAILY_HOURS = 8.0

# Colors
COLOR_HEADER_BG   = "1F3864"   # dark navy
COLOR_MONTH_BG    = "2E75B6"   # medium blue
COLOR_SUBHDR_BG   = "BDD7EE"   # light blue
COLOR_TOTAL_BG    = "FCE4D6"   # light orange
COLOR_ALT_ROW     = "EBF3FB"   # very light blue (alternating rows)
COLOR_WHITE       = "FFFFFF"
COLOR_WORKED      = "D9EAD3"   # light green
COLOR_ABSENCE     = "FCE4D6"   # light red/orange


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
TIME_RE = re.compile(r'\b(\d{2}:\d{2})\b')
DAY_RE  = re.compile(
    r'^(Lu|Ma|Me|Je|Ve|Sa|Di)\s+(\d{2}/\d{2}/\d{4})$'
)


def parse_first_last_times(pointage: str):
    """Return (first_time_str, last_time_str) from a pointage cell, or (None, None)."""
    if not pointage or not isinstance(pointage, str):
        return None, None
    times = TIME_RE.findall(pointage)
    if not times:
        return None, None
    return times[0], times[-1]


def time_to_minutes(t: str) -> int:
    h, m = map(int, t.split(':'))
    return h * 60 + m


def calc_hours(first: str, last: str) -> float:
    """Hours between first entry and last exit (handles cross-midnight)."""
    if not first or not last:
        return 0.0
    f = time_to_minutes(first)
    l = time_to_minutes(last)
    diff = (24 * 60 - f + l) if l < f else (l - f)
    return round(diff / 60, 4)


def is_absence(pointage) -> bool:
    """True if the day has no valid timestamps (pure absence or empty)."""
    if not pointage or not isinstance(pointage, str):
        return True
    return len(TIME_RE.findall(pointage)) == 0


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------
def extract_employees(rows):
    """
    Parse a flat list of row-tuples from the Excel sheet and return a list of
    employee dicts:
      {
        'name': str,
        'matricule': str,
        'service': str,
        'days': [ {'date': datetime, 'month': int, 'hours': float, 'absent': bool} ]
      }
    """
    employees = []
    current_emp = None
    current_service = ""
    in_data = False
    
    print(f"[DEBUG] extract_employees: Processing {len(rows)} rows")

    for idx, row in enumerate(rows):
        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None

        if not c0 and not c1:
            in_data = False
            continue

        # Service line
        if isinstance(c0, str) and c0.startswith("SERVICE / SECTION :"):
            current_service = c0.replace("SERVICE / SECTION :", "").strip()
            print(f"[DEBUG] Found service: {current_service}")
            in_data = False
            continue

        # Employee name line (case-insensitive)
        if isinstance(c0, str) and c0.upper().startswith("NOM"):
            if current_emp is not None:
                employees.append(current_emp)
            # Extract name after "NOM" and optional colon/space
            name_part = c0.split(":", 1)[1].strip() if ":" in c0 else c0[3:].strip()
            current_emp = {
                "name": name_part,
                "matricule": "",
                "service": current_service,
                "days": []
            }
            print(f"[DEBUG] Found employee: {name_part}")
            in_data = False
            continue

        # Matricule line
        if isinstance(c0, str) and c0.startswith("MATRICULE :") and current_emp:
            current_emp["matricule"] = c0.replace("MATRICULE :", "").strip()
            continue

        # Header row — next rows are data
        if isinstance(c0, str) and c0 == "Date":
            print(f"[DEBUG] Found 'Date' header at row {idx}, entering data mode")
            in_data = True
            continue

        # Data row
        if in_data and current_emp and isinstance(c0, str):
            m = DAY_RE.match(c0.strip())
            if m:
                date_str = m.group(2)  # DD/MM/YYYY
                try:
                    dt = datetime.strptime(date_str, "%d/%m/%Y")
                except ValueError:
                    continue

                pointage = c2
                first, last = parse_first_last_times(pointage)
                hours = calc_hours(first, last)
                absent = is_absence(pointage)

                current_emp["days"].append({
                    "date": dt,
                    "month": dt.month,
                    "hours": hours,
                    "absent": absent,
                })
            else:
                # Once we hit non-day rows after data, stop data mode
                in_data = False

    # Don't forget last employee
    if current_emp is not None:
        employees.append(current_emp)
    
    print(f"[DEBUG] Total employees extracted: {len(employees)}")

    return employees


def aggregate_monthly(employees):
    """
    Returns a list of employee summary dicts:
      {
        'name': str, 'matricule': str, 'service': str,
        'monthly': { 1: {'hours': float, 'absent_days': int}, ... 12: {...} }
      }
    """
    result = []
    for emp in employees:
        monthly = {m: {"hours": 0.0, "absent_days": 0} for m in range(1, 13)}
        for day in emp["days"]:
            mo = day["month"]
            if not day["absent"]:
                monthly[mo]["hours"] += day["hours"]
            else:
                monthly[mo]["absent_days"] += 1
        result.append({
            "name": emp["name"],
            "matricule": emp["matricule"],
            "service": emp["service"],
            "monthly": monthly,
        })
    return result


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def bold_font(color="000000", size=10):
    return Font(bold=True, color=color, size=size)


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center")


def apply_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font = font
    if fill:    cell.fill = fill
    if align:   cell.alignment = align
    if border:  cell.border = border
    if num_format: cell.number_format = num_format
    return cell


# ---------------------------------------------------------------------------
# Write pivot sheet
# ---------------------------------------------------------------------------
def write_pivot_sheet(ws, data, sheet_title, year=2025):
    """
    Write the pivot table to the given worksheet.

    Layout:
      Row 1: Title
      Row 2: blank
      Row 3: Header row 1 — employee info cols + month names (merged across H/A sub-cols)
      Row 4: Header row 2 — "Heures" / "Absences" per month + totals
      Row 5+: Data rows
    """

    # ---- helpers for this sheet
    hdr_fill   = hex_fill(COLOR_HEADER_BG)
    month_fill = hex_fill(COLOR_MONTH_BG)
    sub_fill   = hex_fill(COLOR_SUBHDR_BG)
    total_fill = hex_fill(COLOR_TOTAL_BG)
    alt_fill   = hex_fill(COLOR_ALT_ROW)
    w_fill     = hex_fill(COLOR_WORKED)
    a_fill     = hex_fill(COLOR_ABSENCE)
    bdr        = thin_border()
    white_font = Font(color=COLOR_WHITE, bold=True, size=10)
    dark_font  = Font(color="000000", size=9)
    dark_bold  = Font(color="000000", bold=True, size=9)

    # Fixed columns: Matricule, Nom, Service
    FIXED_COLS = 3   # cols A, B, C
    # Each month gets 2 sub-columns: Heures, Absences (jours)
    # Then 2 total columns: Total Heures, Total Absences
    total_cols = FIXED_COLS + 12 * 2 + 2

    # ---- Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    apply_cell(ws, 1, 1,
               f"MEDIDIS — {sheet_title} — Pointage {year}",
               font=Font(bold=True, size=13, color=COLOR_WHITE),
               fill=hdr_fill,
               align=center())

    # ---- Row 3: Month headers (merged over 2 cols each)
    apply_cell(ws, 3, 1, "Matricule", font=white_font, fill=hdr_fill, align=center(), border=bdr)
    apply_cell(ws, 3, 2, "Nom",       font=white_font, fill=hdr_fill, align=center(), border=bdr)
    apply_cell(ws, 3, 3, "Service",   font=white_font, fill=hdr_fill, align=center(), border=bdr)

    for mo in range(1, 13):
        col_start = FIXED_COLS + (mo - 1) * 2 + 1
        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start + 1)
        apply_cell(ws, 3, col_start, MONTHS_FR[mo - 1],
                   font=white_font, fill=month_fill, align=center(), border=bdr)

    # Totals header
    total_h_col = FIXED_COLS + 12 * 2 + 1
    total_a_col = FIXED_COLS + 12 * 2 + 2
    ws.merge_cells(start_row=3, start_column=total_h_col, end_row=3, end_column=total_a_col)
    apply_cell(ws, 3, total_h_col, "TOTAL ANNUEL", font=white_font, fill=hdr_fill, align=center(), border=bdr)

    # ---- Row 4: Sub-headers (Heures / Absences)
    apply_cell(ws, 4, 1, "", fill=hdr_fill, border=bdr)
    apply_cell(ws, 4, 2, "", fill=hdr_fill, border=bdr)
    apply_cell(ws, 4, 3, "", fill=hdr_fill, border=bdr)

    for mo in range(1, 13):
        col_h = FIXED_COLS + (mo - 1) * 2 + 1
        col_a = col_h + 1
        apply_cell(ws, 4, col_h, "Heures",   font=Font(bold=True, size=8, color="1F3864"), fill=sub_fill, align=center(), border=bdr)
        apply_cell(ws, 4, col_a, "Abs (j)",  font=Font(bold=True, size=8, color="7F0000"), fill=sub_fill, align=center(), border=bdr)

    apply_cell(ws, 4, total_h_col, "Heures", font=Font(bold=True, size=9, color="1F3864"), fill=total_fill, align=center(), border=bdr)
    apply_cell(ws, 4, total_a_col, "Abs (j)", font=Font(bold=True, size=9, color="7F0000"), fill=total_fill, align=center(), border=bdr)

    # Freeze panes at row 5
    ws.freeze_panes = ws.cell(row=5, column=4)

    # ---- Data rows
    for idx, emp in enumerate(data):
        row = 5 + idx
        fill_row = alt_fill if idx % 2 == 0 else hex_fill(COLOR_WHITE)

        apply_cell(ws, row, 1, emp["matricule"], font=dark_font, fill=fill_row, align=center(), border=bdr)
        apply_cell(ws, row, 2, emp["name"],      font=dark_bold, fill=fill_row, align=left(),   border=bdr)
        apply_cell(ws, row, 3, emp["service"],   font=dark_font, fill=fill_row, align=left(),   border=bdr)

        total_hours  = 0.0
        total_absent = 0

        for mo in range(1, 13):
            col_h = FIXED_COLS + (mo - 1) * 2 + 1
            col_a = col_h + 1
            h = emp["monthly"][mo]["hours"]
            a = emp["monthly"][mo]["absent_days"]
            total_hours  += h
            total_absent += a

            apply_cell(ws, row, col_h, round(h, 2),
                       font=dark_font, fill=w_fill, align=center(), border=bdr,
                       num_format='0.00')
            apply_cell(ws, row, col_a, a,
                       font=dark_font, fill=a_fill, align=center(), border=bdr)

        apply_cell(ws, row, total_h_col, round(total_hours, 2),
                   font=Font(bold=True, size=9), fill=total_fill, align=center(), border=bdr,
                   num_format='0.00')
        apply_cell(ws, row, total_a_col, total_absent,
                   font=Font(bold=True, size=9), fill=total_fill, align=center(), border=bdr)

    # ---- Column widths
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    for mo in range(1, 13):
        col_h = FIXED_COLS + (mo - 1) * 2 + 1
        ws.column_dimensions[get_column_letter(col_h)].width     = 8
        ws.column_dimensions[get_column_letter(col_h + 1)].width = 7
    ws.column_dimensions[get_column_letter(total_h_col)].width = 10
    ws.column_dimensions[get_column_letter(total_a_col)].width = 9

    # ---- Row heights
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20
    for i in range(len(data)):
        ws.row_dimensions[5 + i].height = 16

    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def get_input_files(input_path):
    """Get list of Excel files from input path (file or directory)."""
    if os.path.isdir(input_path):
        files = []
        for f in os.listdir(input_path):
            if f.lower().endswith('.xlsx') or f.lower().endswith('.xls'):
                files.append(os.path.join(input_path, f))
        return sorted(files)
    elif os.path.isfile(input_path):
        return [input_path]
    else:
        return []


def process_file(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    employees = extract_employees(rows)
    data = aggregate_monthly(employees)
    return data


def generate_output_filename(input_path):
    """Generate output filename based on input filename."""
    base = os.path.splitext(os.path.basename(input_path))[0]
    return f"{base}_pivot.xlsx"


def main():
    parser = argparse.ArgumentParser(description="MEDIDIS Pointage Pivot Generator")
    parser.add_argument("input", nargs="?",
                        default="input",
                        help="Input Excel file or directory (default: input/)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output Excel file (if single input) or directory (if multiple)")
    args = parser.parse_args()

    # Get input files
    input_files = get_input_files(args.input)

    if not input_files:
        print(f"Error: No Excel files found in '{args.input}'")
        sys.exit(1)

    print(f"Found {len(input_files)} input file(s)")

    # Process each file separately
    for path in input_files:
        print(f"\nProcessing: {path}")
        data = process_file(path)
        print(f"  → {len(data)} employees found")

        # Generate output filename
        if args.output:
            if len(input_files) == 1:
                output_path = args.output
            else:
                # Multiple files: use output as directory
                os.makedirs(args.output, exist_ok=True)
                output_path = os.path.join(args.output, generate_output_filename(path))
        else:
            output_path = generate_output_filename(path)

        # Create workbook for this file
        out_wb = Workbook()
        out_wb.remove(out_wb.active)  # remove default blank sheet

        # Sheet name from filename (max 31 chars for Excel)
        base = os.path.splitext(os.path.basename(path))[0][:28]
        ws = out_wb.create_sheet(title=base)
        write_pivot_sheet(ws, data, sheet_title=base)

        out_wb.save(output_path)
        print(f"  → Saved to: {output_path}")

    print(f"\n✓ Processing complete")


def process_annual_pivot(input_dir, output_dir):
    """
    Process annual pivot analysis for production files.
    Returns the output file path or None if no files processed.
    """
    # Get all Excel files from input directory
    input_files = get_input_files(input_dir)
    
    print(f"[DEBUG] Found {len(input_files)} files in input directory")
    
    if not input_files:
        return None
    
    output_files = []
    
    for path in input_files:
        filename = os.path.basename(path)
        print(f"[DEBUG] Processing: {filename}")
        
        data = process_file(path)
        print(f"[DEBUG] Employees found: {len(data) if data else 0}")
        
        if not data:
            continue
        
        # Generate output filename
        base = os.path.splitext(os.path.basename(path))[0]
        output_filename = f"{base}_ANNUAL_PIVOT.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # Create workbook
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        
        # Sheet name from filename (max 31 chars for Excel)
        sheet_name = base[:28]
        ws = out_wb.create_sheet(title=sheet_name)
        write_pivot_sheet(ws, data, sheet_title=sheet_name)
        
        out_wb.save(output_path)
        output_files.append(output_path)
        print(f"[DEBUG] Saved: {output_path}")
    
    # Return the first output file (or None if none created)
    return output_files[0] if output_files else None


if __name__ == "__main__":
    main()
