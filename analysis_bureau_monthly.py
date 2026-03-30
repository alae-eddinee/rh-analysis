import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Suppress warnings from openpyxl if it reads misnamed files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
import os
FOLDER_PATH = os.path.join(os.path.dirname(__file__), "Data")
OUTPUT_FILENAME = "Monthly_Global_Analysis.xlsx"

# LIST OF EMPLOYEES TO EXCLUDE (Case insensitive)
EXCLUDED_EMPLOYEES = [
    # "ABOU HASNAA", 
    "HMOURI ALI"
]

# CODES THAT SIGNIFY AN "OUVRIER" (Worker)
OUVRIER_CODES = ['130', '140', '141', '131']

# Days of week mapping
DAYS_FRENCH = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']

# --- HELPER CLASS FOR XLS COMPATIBILITY ---
class MockCell:
    """Mimics an openpyxl cell object for .xls files read via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Normalizes names to ensure matching works despite spaces/hidden chars."""
    if not name:
        return ""
    name = str(name).upper()
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

def parse_scan_times(scan_str):
    """Parses scan string to count scans and calculate duration."""
    if scan_str is None:
        return [], 0
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    return times, len(times)

def calculate_hours_from_scans(times):
    """Calculates total worked hours from a list of 'HH:MM' strings."""
    if not times:
        return 0.0
    
    total_seconds = 0
    for i in range(0, len(times) - 1, 2):
        try:
            t_in = datetime.strptime(times[i], '%H:%M')
            t_out = datetime.strptime(times[i+1], '%H:%M')
            if t_out < t_in: t_out += timedelta(days=1)
            total_seconds += (t_out - t_in).total_seconds()
        except:
            continue
            
    return round(total_seconds / 3600, 2)

def calculate_lunch_minutes(times):
    """Calculates the duration of the lunch break (gap between scan 2 and 3)."""
    if not times or len(times) < 4:
        return 0
    
    try:
        t_out_lunch = datetime.strptime(times[1], '%H:%M')
        t_in_lunch = datetime.strptime(times[2], '%H:%M')
        
        if t_in_lunch < t_out_lunch: 
            t_in_lunch += timedelta(days=1)
            
        diff_seconds = (t_in_lunch - t_out_lunch).total_seconds()
        return diff_seconds / 60 
    except:
        return 0

def get_sheet_rows(file_path):
    """Generator that yields rows from either .xlsx or .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Warning: '{os.path.basename(file_path)}' is an .xlsx file named as .xls. Switching engine...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Failed to read file with fallback: {e2}")
            else:
                print(f"Error processing .xls file {os.path.basename(file_path)}: {e}")
                return

def process_employee_buffer(employee_data):
    """Decides if an employee is an OUVRIER based on HJ codes."""
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    weekday_recs = [r for r in records if not str(r['day_str']).startswith(('Sa', 'Di'))]
    
    if not weekday_recs:
        return records

    ouvrier_matches = 0
    for r in weekday_recs:
        raw_hj = str(r.get('hj_code', ''))
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in OUVRIER_CODES:
            ouvrier_matches += 1
    
    ratio = ouvrier_matches / len(weekday_recs)
    if ratio > 0.5:
        return []
    
    return records

def extract_month_year_from_filename(file_path):
    """Extracts month and year from filename."""
    filename = os.path.basename(file_path).upper()
    
    # Look for French months in filename
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    # Look for year (4 digits)
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    # Look for month
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    # If no month found, try to find numbers 1-12
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    # Default value
    return '12', year

def extract_date_from_string(date_str):
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(date_str))
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except:
            return None
    return None

def is_ramadan_date(date_obj):
    """Check if a date falls within Ramadan period (Feb 19, 2026 to Mar 20, 2026)."""
    if date_obj is None:
        return False
    ramadan_start = datetime(2026, 2, 19)
    ramadan_end = datetime(2026, 3, 20)
    return ramadan_start <= date_obj <= ramadan_end

def extract_data(file_path):
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    month_num, year_num = extract_month_year_from_filename(file_path)
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''

            if 'SERVICE / SECTION :' in val_0 or val_0.upper().startswith('NOM :'):
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 
                    'matricule': '',
                    'records': []
                }
            elif val_0.upper().startswith('NOM :'):
                raw_name = val_0.split(':', 1)[1].strip() if ':' in val_0 else val_0
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name), 
                    'matricule': '',
                    'records': []
                }
            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
            elif any(val_0.startswith(day) for day in DAYS_FRENCH) and any(char.isdigit() for char in val_0):
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                row_text_upper = (str(val_0) + " " + str(raw_scan_val)).upper()
                date_obj = extract_date_from_string(val_0)
                
                is_leave = 0
                is_holiday = 0
                is_unjustified_absence = 0
                is_day_worked = 0
                hours_worked = 0.0
                daily_target_for_worked_day = 0.0 
                daily_lunch_minutes = 0
                has_lunch_break = 0 
                times_list = []
                scan_count = 0

                is_saturday = val_0.lower().startswith('sa')
                is_sunday = val_0.lower().startswith('di')
                is_ramadan = is_ramadan_date(date_obj)

                if "JOUR FERIE" in row_text_upper:
                    is_holiday = 1
                    if is_sunday: is_holiday = 0 
                elif "CONGE" in row_text_upper:
                    is_leave = 1
                elif "ABSENCE NON JUSTIFIÉE-" in row_text_upper:
                    is_unjustified_absence = 1 
                else:
                    times_list, scan_count = parse_scan_times(raw_scan_val)
                    hours_worked = calculate_hours_from_scans(times_list)
                    
                    if len(times_list) >= 4 and not is_saturday:
                        daily_lunch_minutes = calculate_lunch_minutes(times_list)
                        has_lunch_break = 1
                    
                    if hours_worked > 0:
                        is_day_worked = 1
                        if is_saturday:
                            daily_target_for_worked_day = 4.0
                        elif is_ramadan:
                            daily_target_for_worked_day = 7.0
                        else:
                            daily_target_for_worked_day = 8.0

                if date_obj:
                    day_numeric = date_obj.day
                    day_str = val_0.split()[0] if val_0 else ''

                    record = {
                        'name': current_employee.get('name', ''),
                        'service': current_employee.get('service', ''),
                        'full_date': date_obj, 
                        'day_numeric': day_numeric,
                        'day_str': day_str,
                        'hj_code': str(hj_val).strip(),
                        'times_list': times_list,
                        'hours_worked': hours_worked,
                        'is_day_worked': is_day_worked,
                        'is_leave': is_leave,
                        'is_holiday': is_holiday,
                        'is_unjustified_absence': is_unjustified_absence,
                        'scan_count': scan_count,
                        'daily_target_for_worked_day': daily_target_for_worked_day,
                        'daily_lunch_minutes': daily_lunch_minutes,
                        'has_lunch_break': has_lunch_break,
                        'month_num': month_num,
                        'year_num': year_num
                    }
                    current_employee['records'].append(record)

        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)
    
    except Exception as e:
        print(f"Error opening {os.path.basename(file_path)}: {e}")
        return []
    
    return all_records

def analyze_record(row):
    """Applies business rules to a single daily record."""
    # Initialize Defaults
    is_late_930 = 0
    is_late_1000 = 0
    is_late_1400 = 0
    no_lunch = 0
    is_under = 0
    is_half_day = 0

    if row['is_leave'] or row['is_holiday']:
        return 0, 0, 0, 0, 0, 0

    times = row['times_list']
    if not times: 
        return 0, 0, 0, 0, 0, 0

    # Check if Ramadan period
    is_ramadan = is_ramadan_date(row.get('full_date'))

    first_scan = datetime.strptime(times[0], '%H:%M')
    
    # --- TIME LIMITS ---
    limit_930 = first_scan.replace(hour=9, minute=30, second=0)
    limit_1000 = first_scan.replace(hour=10, minute=0, second=0)
    limit_1300 = first_scan.replace(hour=13, minute=0, second=0)
    limit_1400 = first_scan.replace(hour=14, minute=0, second=0)

    # --- LATENESS LOGIC (ANTI-DUPLICATION) ---
    if first_scan > limit_1400:
        is_late_1400 = 1
        is_late_1000 = 0 
        is_late_930 = 0
    elif first_scan > limit_1000:
        is_late_1400 = 0
        is_late_1000 = 1
        is_late_930 = 0
    elif first_scan > limit_930:
        is_late_1400 = 0
        is_late_1000 = 0
        is_late_930 = 1

    is_saturday = str(row['day_str']).startswith('Sa')
    
    # --- NO LUNCH LOGIC ---
    # During Ramadan, no lunch check - always 0
    if is_ramadan:
        no_lunch = 0
    elif is_late_1400:
        no_lunch = 0
    else:
        no_lunch = 1 if (len(times) < 4 and not is_saturday) else 0

    # --- TARGET HOURS ---
    # During Ramadan: 7h for weekdays, 4h for Saturdays
    if is_saturday:
        target = 4.0
    else:
        target = 7.0 if is_ramadan else 8.0
    
    is_under = 1 if row['hours_worked'] > 0 and row['hours_worked'] < target else 0

    # --- HALF DAY LOGIC (REVISED) ---
    # During Ramadan: threshold is 6.5h (since target is 7h), otherwise 7h (since target is 8h)
    
    if row['is_day_worked'] and not is_saturday and len(times) >= 2:
        try:
            t_entry = datetime.strptime(times[0], '%H:%M')
            t_exit = datetime.strptime(times[-1], '%H:%M')
            if t_exit < t_entry: t_exit += timedelta(days=1)
            
            # --- Condition A: Afternoon Only (Entered after 13:00) ---
            cond_afternoon = (t_entry >= limit_1300)
            
            # --- Condition B: Morning Only (Left before 14:00) ---
            hours_threshold = 6.5 if is_ramadan else 7.0
            cond_morning = (t_exit <= limit_1400) and (row['hours_worked'] < hours_threshold)
            
            if cond_afternoon or cond_morning:
                is_half_day = 1
                
        except:
            pass 

    return is_late_930, is_late_1000, is_late_1400, no_lunch, is_under, is_half_day

def calculate_business_days_in_range(start_date, end_date):
    current = start_date
    business_days = 0
    while current <= end_date:
        wd = current.weekday()
        if wd != 6:
            business_days += 1
        current += timedelta(days=1)
    return business_days

def calculate_weighted_business_days_in_range(start_date, end_date):
    """Calculate weighted business days: weekdays = 1.0, Saturdays = 0.5"""
    current = start_date
    weighted_days = 0
    while current <= end_date:
        wd = current.weekday()
        if wd == 5:  # Saturday
            weighted_days += 0.5
        elif wd != 6:  # Monday-Friday (not Sunday)
            weighted_days += 1.0
        current += timedelta(days=1)
    return weighted_days

def minutes_to_hhmm(mins):
    if pd.isna(mins) or mins == 0:
        return ""
    hours = int(mins // 60)
    minutes = int(round(mins % 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours:02}:{minutes:02}"

def decimal_hours_to_hhmm(decimal_hours):
    if pd.isna(decimal_hours) or decimal_hours == 0:
        return "00:00"
    
    is_negative = decimal_hours < 0
    minutes_total = abs(decimal_hours) * 60
    hours = int(minutes_total // 60)
    minutes = int(round(minutes_total % 60))
    
    if minutes == 60:
        hours += 1
        minutes = 0
        
    time_str = f"{hours:02}:{minutes:02}"
    return f"-{time_str}" if is_negative else time_str

def process_monthly_analysis(input_dir, output_dir):
    """
    Traite les fichiers dans input_dir et sauvegarde l'analyse mensuelle dans output_dir.
    Retourne le chemin du fichier généré ou None.
    """
    if not os.path.exists(input_dir):
        print(f"Dossier non trouvé : {input_dir}")
        return None

    # S'assurer que le dossier de sortie existe
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []
    print("Reading files...")
    for file in os.listdir(input_dir):
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Processing: {file}")
            all_data.extend(extract_data(os.path.join(input_dir, file)))

    if not all_data:
        print("No data found.")
        return None

    df = pd.DataFrame(all_data)

    # --- DÉTECTION CHRONOLOGIQUE AMÉLIORÉE ---
    if 'day_numeric' in df.columns and not df.empty:
        # 1. Récupérer les infos de base
        month_num = df['month_num'].iloc[0] if 'month_num' in df.columns else '01'
        year_num = df['year_num'].iloc[0] if 'year_num' in df.columns else '2026'
        
        # 2. Identifier la séquence chronologique réelle
        unique_days_in_order = []
        seen = set()
        for d in df['day_numeric']:
            if d not in seen:
                unique_days_in_order.append(d)
                seen.add(d)

        real_start_day = unique_days_in_order[0]
        real_end_day = unique_days_in_order[-1]
        
        # Détecter s'il y a une transition de mois (ex: 25, 26... 31, 1, 2)
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days_in_order) - 1):
            if unique_days_in_order[i] > unique_days_in_order[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        print(f"\n--- ANALYSE DE LA PÉRIODE ---")
        print(f"Séquence détectée : {unique_days_in_order}")
        
        # 3. Définir le jour cible (le dernier jour chronologique)
        target_report_day = real_end_day
        
        # 4. Vérifier si le dernier jour est complet (Scan count)
        last_day_records = df[df['day_numeric'] == target_report_day]
        total_last_day = len(last_day_records)
        # On considère un jour incomplet si + de 50% des gens n'ont qu'un seul pointage (ou 0)
        incomplete_count = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total_last_day > 0 and (incomplete_count / total_last_day) > 0.5:
            print(f"DÉCISION : Le jour {target_report_day} est incomplet (en cours).")
            # Supprimer le jour incomplet du DataFrame pour l'analyse
            df = df[df['day_numeric'] != target_report_day].copy()
            # Le nouveau jour cible devient le précédent dans la liste ordonnée
            if len(unique_days_in_order) > 1:
                target_report_day = unique_days_in_order[-2]
                real_end_day = target_report_day
            print(f"Nouveau jour cible : {target_report_day}")
        else:
            print(f"DÉCISION : Le jour {target_report_day} est complet.")

        # 5. Calcul du nom du mois pour le header
        month_names = {
            '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
            '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
            '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
        }
        month_name = month_names.get(month_num, f'Mois {month_num}')
        
        # 6. Créer les dates complètes pour le calcul des jours ouvrés
        if 'full_date' in df.columns and not df['full_date'].isnull().all():
            # Utiliser les dates réelles si disponibles
            final_min_date = df['full_date'].min()
            final_max_date = df['full_date'].max()
        else:
            # Recréer les dates à partir des informations extraites
            final_min_date = datetime(int(year_num), int(month_num), real_start_day)
            final_max_date = datetime(int(year_num), int(month_num), real_end_day)
            
            # Gérer les périodes multi-mois
            if has_transition:
                # Si transition, le dernier mois est probablement le mois suivant
                if month_num == '12':
                    next_month_num = '01'
                    next_year_num = str(int(year_num) + 1)
                else:
                    next_month_num = f"{int(month_num) + 1:02d}"
                    next_year_num = year_num
                final_max_date = datetime(int(next_year_num), int(next_month_num), real_end_day)
        
        print(f"\n--- PLAGE DE JOURS DÉTECTÉE ---")
        print(f"Premier jour trouvé : {real_start_day}")
        print(f"Dernier jour trouvé : {real_end_day}")
        
        # Calculer correctement le total de jours pour les périodes multi-mois
        if has_transition:
            # Période multi-mois : jours du premier mois + jours du deuxième mois
            first_month_days = unique_days_in_order[:pivot_index + 1]
            second_month_days = unique_days_in_order[pivot_index + 1:]
            total_days = len(first_month_days) + len(second_month_days)
            print(f"Période multi-mois détectée : {len(first_month_days)} jours + {len(second_month_days)} jours")
        else:
            # Période simple
            total_days = len(unique_days_in_order)
        
        print(f"Total jours analysés : {total_days}")
        print(f"Final Analysis Period: {final_min_date.strftime('%d/%m/%Y')} to {final_max_date.strftime('%d/%m/%Y')}")
        global_expected_days = calculate_business_days_in_range(final_min_date, final_max_date)
        print(f"Theoretical Business Days (Mon-Sat) in period: {global_expected_days}")
        
        # Créer un nom de fichier dynamique basé sur la période analysée
        dynamic_filename = f"Monthly_Global_Analysis_{real_start_day:02d}-{month_num}-{year_num}_A_{real_end_day:02d}-{month_num}-{year_num}.xlsx"
        output_path = os.path.join(output_dir, dynamic_filename)
        
        # Use actual calculated dates for the header to handle multi-month periods correctly
        header_start = final_min_date.strftime('%d/%m/%Y')
        header_end = final_max_date.strftime('%d/%m/%Y')
        header_text = f"Analyse Mensuelle - Période : {header_start} au {header_end}"

    else:
        print("Could not detect valid dates. Exiting.")
        return None

    if EXCLUDED_EMPLOYEES:
        print(f"\nFiltering out: {EXCLUDED_EMPLOYEES}")
        excluded_clean = [clean_name_string(name) for name in EXCLUDED_EMPLOYEES]
        df = df[~df['name'].isin(excluded_clean)]

    if df.empty:
        print("All data filtered out.")
        return None

    print("Analyzing metrics...")
    metrics = df.apply(analyze_record, axis=1)
    
    df['ENTRY > 9H30'] = [x[0] for x in metrics]
    df['ENTRY > 10H'] = [x[1] for x in metrics]
    df['ENTRY > 14H'] = [x[2] for x in metrics] 
    df['NO LUNCH'] = [x[3] for x in metrics]
    df['UNDER 8H'] = [x[4] for x in metrics]
    df['IS HALF DAY'] = [x[5] for x in metrics]

    # Split data into Ramadan and Normal days
    df['is_ramadan'] = df['full_date'].apply(is_ramadan_date)
    ramadan_df = df[df['is_ramadan'] == True].copy()
    normal_df = df[df['is_ramadan'] == False].copy()
    has_ramadan_data = len(ramadan_df) > 0
    has_normal_data = len(normal_df) > 0

    # Calculate date ranges and working days for each table
    normal_date_range = None
    ramadan_date_range = None
    normal_working_days = 0
    ramadan_working_days = 0
    
    if has_normal_data:
        normal_min_date = normal_df['full_date'].min()
        normal_max_date = normal_df['full_date'].max()
        normal_working_days = calculate_weighted_business_days_in_range(normal_min_date, normal_max_date)
        normal_date_range = (normal_min_date, normal_max_date)
        print(f"DEBUG: Normal table date range: {normal_min_date} to {normal_max_date}")
        print(f"DEBUG: Normal table weighted working days: {normal_working_days}")
    
    if has_ramadan_data:
        ramadan_min_date = ramadan_df['full_date'].min()
        ramadan_max_date = ramadan_df['full_date'].max()
        ramadan_working_days = calculate_weighted_business_days_in_range(ramadan_min_date, ramadan_max_date)
        ramadan_date_range = (ramadan_min_date, ramadan_max_date)
        print(f"DEBUG: Ramadan table date range: {ramadan_min_date} to {ramadan_max_date}")
        print(f"DEBUG: Ramadan table weighted working days: {ramadan_working_days}")

    # Function to generate report from a dataframe subset
    def generate_report_from_df(subset_df, table_expected_days, is_ramadan_table=False):
        if subset_df.empty:
            return None, None
        
        # Calculate weighted days worked: weekdays = 1.0, Saturdays = 0.5
        def calc_weighted_days_worked(group):
            weekdays = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_day_worked'] == 1)]
            saturdays = group[(group['day_str'].str.startswith('Sa', na=False)) & (group['is_day_worked'] == 1)]
            # Count unique days for each
            weekday_days = len(weekdays['day_numeric'].unique())
            saturday_days = len(saturdays['day_numeric'].unique())
            return weekday_days + (saturday_days * 0.5)
        
        # Calculate weighted absence: weekday absences = 1.0, Saturday absences = 0.5
        def calc_weighted_absence(group):
            weekday_absences = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_unjustified_absence'] == 1)]
            saturday_absences = group[(group['day_str'].str.startswith('Sa', na=False)) & (group['is_unjustified_absence'] == 1)]
            # Count unique days for each
            weekday_abs_count = len(weekday_absences['day_numeric'].unique())
            saturday_abs_count = len(saturday_absences['day_numeric'].unique())
            return weekday_abs_count + (saturday_abs_count * 0.5)
        
        # Calculate weighted days worked per employee
        weighted_days = subset_df.groupby('name').apply(calc_weighted_days_worked).reset_index()
        weighted_days.columns = ['name', 'weighted_days_worked']
        
        # Calculate weighted absence per employee
        weighted_absence = subset_df.groupby('name').apply(calc_weighted_absence).reset_index()
        weighted_absence.columns = ['name', 'weighted_absence']
        
        report = subset_df.groupby('name').agg({
            'is_day_worked': 'sum',
            'is_leave': 'sum',
            'is_holiday': 'sum',
            'is_unjustified_absence': 'sum',
            'daily_target_for_worked_day': 'sum', 
            'ENTRY > 10H': 'sum',
            'ENTRY > 14H': 'sum', 
            'ENTRY > 9H30': 'sum',
            'NO LUNCH': 'sum',
            'UNDER 8H': 'sum',
            'IS HALF DAY': 'sum',
            'hours_worked': 'sum',
            'daily_lunch_minutes': 'sum',
            'has_lunch_break': 'sum'
        }).reset_index()
        
        # Merge weighted days worked
        report = report.merge(weighted_days, on='name', how='left')
        
        # Merge weighted absence (overrides the simple sum)
        report = report.merge(weighted_absence, on='name', how='left')
        
        report.rename(columns={
            'name': 'Employee name',
            'weighted_days_worked': 'days worked',
            'weighted_absence': 'ABSENCE',
            'daily_target_for_worked_day': 'TOTAL HOURS NEEDED', 
            'hours_worked': 'TOTAL HOURS WORKED',
            'IS HALF DAY': 'HALF DAYS'
        }, inplace=True)
        # Rename UNDER 8H based on table type
        if is_ramadan_table:
            report.rename(columns={'UNDER 8H': 'UNDER 7H'}, inplace=True)
        saturday_records = subset_df[subset_df['day_str'].str.startswith('Sa', na=False)]
        expected_saturdays = len(saturday_records['day_numeric'].unique())
        # Use table-specific working days, not global
        report['real working days'] = table_expected_days - report['is_leave'] - report['is_holiday']
        saturday_work_by_employee = saturday_records[saturday_records['is_day_worked'] == 1].groupby('name').size().reset_index(name='saturdays_worked')
        saturday_work_by_employee['saturdays_worked'] = saturday_work_by_employee['saturdays_worked'].apply(lambda x: 1 if x > 0 else 0)
        report = report.merge(saturday_work_by_employee, left_on='Employee name', right_on='name', how='left')
        report['saturdays_worked'] = report['saturdays_worked'].fillna(0)
        if 'name' in report.columns:
            report.drop('name', axis=1, inplace=True)
        report['saturdays_absent'] = expected_saturdays - report['saturdays_worked']
        report['weekdays_absent'] = report['real working days'] - report['saturdays_absent'] - report['days worked']
        # Only count unjustified absences, not calculated absences
        report.drop(['saturdays_worked', 'saturdays_absent', 'weekdays_absent'], axis=1, inplace=True)
        report['avg_lunch_raw'] = report.apply(
            lambda x: x['daily_lunch_minutes'] / x['has_lunch_break'] if x['has_lunch_break'] > 0 else (x['daily_lunch_minutes'] if x['daily_lunch_minutes'] > 0 else 0), axis=1
        )
        report['AVG LUNCH TIME'] = report['avg_lunch_raw'].apply(minutes_to_hhmm)
        # Calculate balance BEFORE converting to HH:MM format
        report['balance_raw'] = report['TOTAL HOURS WORKED'] - report['TOTAL HOURS NEEDED']
        # Convert hours columns to HH:MM format
        report['TOTAL HOURS NEEDED'] = report['TOTAL HOURS NEEDED'].apply(decimal_hours_to_hhmm)
        report['TOTAL HOURS WORKED'] = report['TOTAL HOURS WORKED'].apply(decimal_hours_to_hhmm)
        report['Balance of hours worked'] = report['balance_raw'].apply(decimal_hours_to_hhmm)
        # Get column list and mark which ones should show '-' for Ramadan
        final_cols = [
            'Employee name', 
            'real working days', 
            'days worked',
            'ABSENCE', 
            'HALF DAYS', 
            'UNDER 8H' if not is_ramadan_table else 'UNDER 7H', 
            'NO LUNCH', 
            'AVG LUNCH TIME',
            'ENTRY > 14H', 
            'ENTRY > 10H', 
            'ENTRY > 9H30', 
            'TOTAL HOURS NEEDED', 
            'TOTAL HOURS WORKED', 
            'Balance of hours worked'
        ]
        # Columns to show '-' for Ramadan
        ramadan_dash_cols = ['NO LUNCH', 'AVG LUNCH TIME', 'ENTRY > 14H']
        return report[final_cols].sort_values('ABSENCE', ascending=False), (ramadan_dash_cols if is_ramadan_table else [])

    # Generate reports with table-specific working days
    normal_report, _ = generate_report_from_df(normal_df, normal_working_days, is_ramadan_table=False) if has_normal_data else (None, [])
    ramadan_report, ramadan_dash_cols = generate_report_from_df(ramadan_df, ramadan_working_days, is_ramadan_table=True) if has_ramadan_data else (None, [])

    # --- EXPORT ---
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('Monthly Summary')
            
            # Formats
            header_title = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 14, 'font_color': '#2F5597', 'border': 1
            })
            section_header = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 12, 'fg_color': '#E7E6E6', 'border': 1
            })
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            header_red = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#C00000', 'font_color': 'white', 'border': 1
            })
            header_orange = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#ED7D31', 'font_color': 'white', 'border': 1
            })
            body_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
            text_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'})
            absence_red_format = workbook.add_format({
                'border': 1, 'align': 'center', 'valign': 'vcenter', 
                'bg_color': '#FFC7CE', 'font_color': '#9C0006'
            })
            
            current_row = 0
            num_cols = 15
            
            # Write main title
            worksheet.merge_range(0, 0, 0, num_cols - 1, header_text, header_title)
            current_row = 2
            
            # Helper function to write a table
            def write_table(report_df, table_title, start_row, dash_columns=None):
                if report_df is None or report_df.empty:
                    return start_row
                dash_columns = dash_columns or []
                # Write section header
                worksheet.merge_range(start_row, 0, start_row, num_cols - 1, table_title, section_header)
                start_row += 1
                # Write column headers
                for col_num, value in enumerate(report_df.columns.values):
                    if "14H" in str(value):
                        worksheet.write(start_row, col_num, value, header_red)
                    elif "HALF DAYS" in str(value):
                        worksheet.write(start_row, col_num, value, header_orange)
                    else:
                        worksheet.write(start_row, col_num, value, header_format)
                start_row += 1
                # Write data rows
                for row_idx, row_data in report_df.iterrows():
                    for col_num, col_name in enumerate(report_df.columns):
                        value = row_data[col_name]
                        # Show '-' for Ramadan dash columns
                        if col_name in dash_columns:
                            worksheet.write(start_row, col_num, "-", body_format)
                            continue
                        if pd.isna(value): 
                            value = ""
                        # Show zeros for count columns, empty strings for time columns
                        if col_name in ['real working days', 'days worked', 'ABSENCE', 'HALF DAYS', 'UNDER 8H', 'UNDER 7H', 'NO LUNCH', 'ENTRY > 14H', 'ENTRY > 10H', 'ENTRY > 9H30']:
                            if value == 0: 
                                value = 0
                        elif col_name in ['AVG LUNCH TIME', 'Balance of hours worked', 'TOTAL HOURS WORKED']:
                            if value == 0 or value == "00:00": 
                                value = ""
                        # Determine cell format
                        if col_name == 'Employee name':
                            cell_fmt = text_format
                        else:
                            cell_fmt = body_format
                        # Apply red formatting for ABSENCE column if value > 0
                        if col_name == 'ABSENCE' and value > 0:
                            worksheet.write(start_row, col_num, value, absence_red_format)
                        else:
                            worksheet.write(start_row, col_num, value, cell_fmt)
                    start_row += 1
                return start_row + 1  # Add spacing after table
            
            # Write Normal Days table first (if exists)
            if has_normal_data:
                normal_start = normal_date_range[0].strftime('%d/%m/%Y') if normal_date_range else ''
                normal_end = normal_date_range[1].strftime('%d/%m/%Y') if normal_date_range else ''
                normal_title = f"JOURS NORMAUX (8h/jour) - Du {normal_start} au {normal_end}"
                current_row = write_table(normal_report, normal_title, current_row, dash_columns=[])
            
            # Write Ramadan Days table (if exists)
            if has_ramadan_data:
                ramadan_start = ramadan_date_range[0].strftime('%d/%m/%Y') if ramadan_date_range else '19/02/2026'
                ramadan_end = ramadan_date_range[1].strftime('%d/%m/%Y') if ramadan_date_range else '20/03/2026'
                ramadan_title = f"JOURS DU RAMADAN (7h/jour, pas de déjeuner) - Du {ramadan_start} au {ramadan_end}"
                current_row = write_table(ramadan_report, ramadan_title, current_row, dash_columns=ramadan_dash_cols)
            
            # Set column widths
            for i in range(num_cols):
                if i == 0:  # Employee name
                    worksheet.set_column(i, i, 20)
                elif i in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:  # Count columns
                    worksheet.set_column(i, i, 10)
                elif i in [11, 12, 13]:  # Time/Hour columns
                    worksheet.set_column(i, i, 14)
                else:
                    worksheet.set_column(i, i, 12)

        print(f"\nSUCCESS! Monthly report generated: {output_path}")
        return output_path

    except Exception as e:
        print(f"Error saving file: {e}")
        import traceback
        print(f"DEBUG: {traceback.format_exc()}")
        return None

def main():
    if not os.path.exists(FOLDER_PATH):
        print("Folder not found.")
        return

    output = process_monthly_analysis(FOLDER_PATH, FOLDER_PATH)
    if output:
        print(f"Report generated: {output}")

if __name__ == "__main__":
    main()