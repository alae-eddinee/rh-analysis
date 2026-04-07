import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Suppress warnings from openpyxl if it reads misnamed files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
FOLDER_PATH = os.path.join(os.path.dirname(__file__), "Data")
OUTPUT_FILENAME = "Monthly_Production_Analysis.xlsx"

# LIST OF EMPLOYEES TO EXCLUDE (Case insensitive)
EXCLUDED_EMPLOYEES = [
    # "ABOU HASNAA", 
    "HMOURI ALI"
]

# PRODUCTION WORKER CODES - Only process employees with these codes
PRODUCTION_CODES = ['130', '131', '140', '141']

# Hours per day for production workers
WEEKDAY_HOURS = 9.0
SATURDAY_HOURS = 5.0

# Days of week mapping
DAYS_FRENCH = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']

# --- HELPER CLASS FOR XLS COMPATIBILITY ---
class MockCell:
    """Mimics an openpyxl cell object for .xls files read via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Minimal cleaning - only remove hidden chars, preserve original case and spacing."""
    if not name:
        return ""
    name = str(name)
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    return name.strip()

def parse_scan_times(scan_str):
    """Parses scan string to count scans."""
    if scan_str is None:
        return [], 0
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    return times, len(times)

def parse_tps_eff(tps_eff_val):
    """
    Parses the 'Tps Eff' column (col 4) from the pointage file.
    Returns hours as float, or None if unparseable/zero.
    Using Tps Eff ensures accurate hours matching the source system,
    which auto-deducts lunch for certain contract types (e.g. HJ 140).
    """
    if tps_eff_val is None:
        return None
    s = str(tps_eff_val).strip()
    m = re.match(r'^(\d+):(\d{2})$', s)
    if m:
        val = int(m.group(1)) + int(m.group(2)) / 60.0
        return val if val > 0 else None
    try:
        f = float(s)
        if 0.0 < f < 5.0:
            return round(f * 24, 4)
    except ValueError:
        pass
    return None

def calculate_lunch_minutes(times):
    """Calculates the duration of the lunch break (gap between scan 2 and 3)."""
    if not times or len(times) < 4:
        return 0
    try:
        t_out_lunch = datetime.strptime(times[1], '%H:%M')
        t_in_lunch  = datetime.strptime(times[2], '%H:%M')
        if t_in_lunch < t_out_lunch: 
            t_in_lunch += timedelta(days=1)
        diff_seconds = (t_in_lunch - t_out_lunch).total_seconds()
        return diff_seconds / 60 
    except:
        return 0

def get_sheet_rows(file_path):
    """Generator that yields rows from either .xlsx or .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Warning: '{os.path.basename(file_path)}' is an .xlsx file named as .xls. Switching engine...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Failed to read file with fallback: {e2}")
            else:
                print(f"Error processing .xls file {os.path.basename(file_path)}: {e}")
                return

def process_employee_buffer(employee_data):
    """Decides if an employee is a Production worker based on HJ codes."""
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    weekday_recs = [r for r in records if not str(r['day_str']).startswith(('Sa', 'Di'))]
    
    if not weekday_recs:
        return records

    prod_matches = 0
    for r in weekday_recs:
        raw_hj = str(r.get('hj_code', ''))
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
        if hj in PRODUCTION_CODES:
            prod_matches += 1
    
    ratio = prod_matches / len(weekday_recs)
    # Production analysis: ONLY include if more than 50% of codes are production codes
    if ratio > 0.5:
        return records
    
    return []

def extract_month_year_from_filename(file_path):
    """Extracts month and year from filename."""
    filename = os.path.basename(file_path).upper()
    
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    return '12', year

def extract_date_from_string(date_str):
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(date_str))
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except:
            return None
    return None

def is_friday_date(date_obj):
    """Check if date is Friday (weekday 4)."""
    if date_obj is None:
        return False
    return date_obj.weekday() == 4

def extract_data(file_path):
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    month_num, year_num = extract_month_year_from_filename(file_path)
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''

            if 'SERVICE / SECTION :' in val_0 or val_0.upper().startswith('NOM :'):
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 
                    'matricule': '',
                    'records': []
                }
            elif val_0.upper().startswith('NOM :'):
                raw_name = val_0.split(':', 1)[1].strip() if ':' in val_0 else val_0
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name), 
                    'matricule': '',
                    'records': []
                }
            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
            elif any(val_0.startswith(day) for day in DAYS_FRENCH) and any(char.isdigit() for char in val_0):
                hj_val      = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                # FIX: Read Tps Eff (col4) for accurate hours
                tps_eff_val = row[4].value if len(row) > 4 else None

                row_text_upper = (str(val_0) + " " + str(raw_scan_val)).upper()
                date_obj = extract_date_from_string(val_0)
                
                is_leave = 0
                is_holiday = 0
                is_unjustified_absence = 0
                is_day_worked = 0
                hours_worked = 0.0
                daily_target_for_worked_day = 0.0 
                daily_lunch_minutes = 0
                has_lunch_break = 0 
                times_list = []
                scan_count = 0

                is_sat = val_0.lower().startswith('sa')
                is_fri = is_friday_date(date_obj)

                if "JOUR FERIE" in row_text_upper:
                    is_holiday = 1
                elif "CONGE" in row_text_upper:
                    is_leave = 1
                elif "ABSENCE AUTORIS" in row_text_upper:
                    # FIX: Treat authorized absence as leave
                    is_leave = 1
                elif "ABSENCE NON JUSTIFI" in row_text_upper:
                    is_unjustified_absence = 1
                else:
                    times_list, scan_count = parse_scan_times(raw_scan_val)

                    # FIX: Use Tps Eff for accurate hours; fallback to raw scan calculation
                    tps_eff_hours = parse_tps_eff(tps_eff_val)
                    if tps_eff_hours is not None:
                        hours_worked = round(tps_eff_hours, 2)
                    else:
                        total_sec = 0
                        for i in range(0, len(times_list) - 1, 2):
                            try:
                                t_in  = datetime.strptime(times_list[i], '%H:%M')
                                t_out = datetime.strptime(times_list[i+1], '%H:%M')
                                if t_out < t_in: t_out += timedelta(days=1)
                                total_sec += (t_out - t_in).total_seconds()
                            except:
                                continue
                        hours_worked = round(total_sec / 3600, 2)
                    
                    if len(times_list) >= 4 and not is_sat:
                        daily_lunch_minutes = calculate_lunch_minutes(times_list)
                        has_lunch_break = 1
                    
                    if hours_worked > 0:
                        is_day_worked = 1
                        daily_target_for_worked_day = SATURDAY_HOURS if is_sat else WEEKDAY_HOURS

                if date_obj:
                    day_numeric = date_obj.day
                    day_str = val_0.split()[0] if val_0 else ''

                    record = {
                        'name': current_employee.get('name', ''),
                        'service': current_employee.get('service', ''),
                        'full_date': date_obj, 
                        'day_numeric': day_numeric,
                        'day_str': day_str,
                        'hj_code': str(hj_val).strip(),
                        'times_list': times_list,
                        'hours_worked': hours_worked,
                        'is_day_worked': is_day_worked,
                        'is_leave': is_leave,
                        'is_holiday': is_holiday,
                        'is_unjustified_absence': is_unjustified_absence,
                        'scan_count': scan_count,
                        'daily_target_for_worked_day': daily_target_for_worked_day,
                        'daily_lunch_minutes': daily_lunch_minutes,
                        'has_lunch_break': has_lunch_break,
                        'month_num': month_num,
                        'year_num': year_num,
                        'is_friday': is_fri,
                        'is_saturday': is_sat
                    }
                    current_employee['records'].append(record)

        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)
    
    except Exception as e:
        print(f"Error opening {os.path.basename(file_path)}: {e}")
        import traceback
        print(traceback.format_exc())
        return []
    
    return all_records

def analyze_record(row):
    """Applies business rules to a single daily record for Production workers."""
    is_late_800  = 0
    is_late_1000 = 0
    no_lunch  = 0
    is_under  = 0
    is_half_day = 0

    if row['is_leave'] or row['is_holiday']:
        return 0, 0, 0, 0, 0

    times = row['times_list']
    if not times: 
        return 0, 0, 0, 0, 0

    first_scan = datetime.strptime(times[0], '%H:%M')
    
    limit_800  = first_scan.replace(hour=8,  minute=0, second=0)
    limit_1000 = first_scan.replace(hour=10, minute=0, second=0)
    limit_1300 = first_scan.replace(hour=13, minute=0, second=0)

    if first_scan > limit_1000:
        is_late_1000 = 1
        is_late_800  = 0
    elif first_scan > limit_800:
        is_late_1000 = 0
        is_late_800  = 1

    is_sat = row.get('is_saturday', False)
    
    if is_sat:
        no_lunch = 0
    else:
        no_lunch = 1 if len(times) < 4 else 0

    target = SATURDAY_HOURS if is_sat else WEEKDAY_HOURS
    is_under = 1 if row['hours_worked'] > 0 and row['hours_worked'] < target else 0

    # Half day: entered after 13:00 OR left before 13:00 with < 8h worked
    if row['is_day_worked'] and not is_sat and len(times) >= 2:
        try:
            t_entry = datetime.strptime(times[0], '%H:%M')
            t_exit  = datetime.strptime(times[-1], '%H:%M')
            if t_exit < t_entry: t_exit += timedelta(days=1)
            
            cond_afternoon = (t_entry >= limit_1300)
            cond_morning   = (t_exit <= limit_1300) and (row['hours_worked'] < 8.0)
            
            if cond_afternoon or cond_morning:
                is_half_day = 1
        except:
            pass 

    return is_late_800, is_late_1000, no_lunch, is_under, is_half_day

def calculate_weighted_business_days_in_range(start_date, end_date):
    """Calculate weighted business days: weekdays = 1.0, Saturdays = 0.5"""
    current = start_date
    weighted_days = 0.0
    while current <= end_date:
        wd = current.weekday()
        if wd == 5:
            weighted_days += 0.5
        elif wd != 6:
            weighted_days += 1.0
        current += timedelta(days=1)
    return weighted_days

def minutes_to_hhmm(mins):
    if pd.isna(mins) or mins == 0:
        return ""
    hours   = int(mins // 60)
    minutes = int(round(mins % 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours:02}:{minutes:02}"

def decimal_hours_to_hhmm(decimal_hours):
    if pd.isna(decimal_hours) or decimal_hours == 0:
        return "00:00"
    is_negative   = decimal_hours < 0
    minutes_total = abs(decimal_hours) * 60
    hours   = int(minutes_total // 60)
    minutes = int(round(minutes_total % 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    time_str = f"{hours:02}:{minutes:02}"
    return f"-{time_str}" if is_negative else time_str

def process_production_monthly_analysis(input_dir, output_dir):
    """Process Production worker files and generate monthly analysis."""
    if not os.path.exists(input_dir):
        print(f"Folder not found: {input_dir}")
        return None

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []
    print("Reading Production files...")
    for file in os.listdir(input_dir):
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Processing: {file}")
            all_data.extend(extract_data(os.path.join(input_dir, file)))

    if not all_data:
        print("No Production data found.")
        return None

    df = pd.DataFrame(all_data)

    # --- CHRONOLOGICAL DETECTION ---
    if 'day_numeric' in df.columns and not df.empty:
        month_num = df['month_num'].iloc[0] if 'month_num' in df.columns else '01'
        year_num  = df['year_num'].iloc[0]  if 'year_num'  in df.columns else '2026'
        
        unique_days_in_order = []
        seen = set()
        for d in df['day_numeric']:
            if d not in seen:
                unique_days_in_order.append(d)
                seen.add(d)

        real_start_day = unique_days_in_order[0]
        real_end_day   = unique_days_in_order[-1]
        
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days_in_order) - 1):
            if unique_days_in_order[i] > unique_days_in_order[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        print(f"\n--- PERIOD ANALYSIS ---")
        print(f"Detected sequence: {unique_days_in_order}")
        
        target_report_day = real_end_day
        
        last_day_records = df[df['day_numeric'] == target_report_day]
        total_last_day   = len(last_day_records)
        incomplete_count = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total_last_day > 0 and (incomplete_count / total_last_day) > 0.5:
            print(f"DECISION: Day {target_report_day} is incomplete (in progress).")
            df = df[df['day_numeric'] != target_report_day].copy()
            if len(unique_days_in_order) > 1:
                target_report_day = unique_days_in_order[-2]
                real_end_day = target_report_day
            print(f"New target day: {target_report_day}")
        else:
            print(f"DECISION: Day {target_report_day} is complete.")

        month_names = {
            '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
            '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
            '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
        }
        month_name = month_names.get(month_num, f'Mois {month_num}')
        
        if 'full_date' in df.columns and not df['full_date'].isnull().all():
            final_min_date = df['full_date'].min()
            final_max_date = df['full_date'].max()
        else:
            final_min_date = datetime(int(year_num), int(month_num), real_start_day)
            final_max_date = datetime(int(year_num), int(month_num), real_end_day)
            if has_transition:
                if month_num == '12':
                    next_month_num = '01'
                    next_year_num  = str(int(year_num) + 1)
                else:
                    next_month_num = f"{int(month_num) + 1:02d}"
                    next_year_num  = year_num
                final_max_date = datetime(int(next_year_num), int(next_month_num), real_end_day)
        
        print(f"\n--- DATE RANGE DETECTED ---")
        print(f"First day found: {real_start_day}")
        print(f"Last day found: {real_end_day}")
        
        if has_transition:
            first_month_days  = unique_days_in_order[:pivot_index + 1]
            second_month_days = unique_days_in_order[pivot_index + 1:]
            total_days = len(first_month_days) + len(second_month_days)
            print(f"Multi-month period: {len(first_month_days)} + {len(second_month_days)} days")
        else:
            total_days = len(unique_days_in_order)
        
        print(f"Total days analyzed: {total_days}")
        global_expected_days = calculate_weighted_business_days_in_range(final_min_date, final_max_date)
        print(f"Theoretical Business Days (weighted): {global_expected_days}")
        
        dynamic_filename = f"Monthly_Production_Analysis_{real_start_day:02d}-{month_num}-{year_num}_A_{real_end_day:02d}-{month_num}-{year_num}.xlsx"
        output_path  = os.path.join(output_dir, dynamic_filename)
        header_text  = f"Analyse Mensuelle Production - Période : {final_min_date.strftime('%d/%m/%Y')} au {final_max_date.strftime('%d/%m/%Y')}"

    else:
        print("Could not detect valid dates. Exiting.")
        return None

    if EXCLUDED_EMPLOYEES:
        print(f"\nFiltering out: {EXCLUDED_EMPLOYEES}")
        excluded_clean = [clean_name_string(name) for name in EXCLUDED_EMPLOYEES]
        df = df[~df['name'].isin(excluded_clean)]

    if df.empty:
        print("All data filtered out.")
        return None

    print("Analyzing Production metrics...")
    metrics = df.apply(analyze_record, axis=1)
    
    df['ENTRY > 8H']  = [x[0] for x in metrics]
    df['ENTRY > 10H'] = [x[1] for x in metrics]
    df['NO LUNCH']    = [x[2] for x in metrics]
    df['UNDER 9H']    = [x[3] for x in metrics]
    df['IS HALF DAY'] = [x[4] for x in metrics]

    min_date = df['full_date'].min()
    max_date = df['full_date'].max()
    working_days = calculate_weighted_business_days_in_range(min_date, max_date)
    
    print(f"DEBUG: Production date range: {min_date} to {max_date}")
    print(f"DEBUG: Production weighted working days: {working_days}")

    def generate_production_report(subset_df, table_expected_days):
        if subset_df.empty:
            return None
        
        # First pointage date per employee
        starting_days = subset_df.groupby('name')['full_date'].min().reset_index()
        starting_days.columns = ['name', 'starting_day']
        
        def calc_weighted_days_worked(group):
            weekdays  = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_day_worked'] == 1)]
            saturdays = group[(group['day_str'].str.startswith('Sa', na=False))  & (group['is_day_worked'] == 1)]
            return len(weekdays['day_numeric'].unique()) + len(saturdays['day_numeric'].unique()) * 0.5
        
        def calc_weighted_absence(group):
            weekday_abs  = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_unjustified_absence'] == 1)]
            saturday_abs = group[(group['day_str'].str.startswith('Sa', na=False))  & (group['is_unjustified_absence'] == 1)]
            return len(weekday_abs['day_numeric'].unique()) + len(saturday_abs['day_numeric'].unique()) * 0.5

        # FIX: Weighted leave/holiday (Saturday = 0.5)
        def calc_weighted_leave(group):
            weekday_lv  = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_leave'] == 1)]
            saturday_lv = group[(group['day_str'].str.startswith('Sa', na=False))  & (group['is_leave'] == 1)]
            return len(weekday_lv) * 1.0 + len(saturday_lv) * 0.5

        def calc_weighted_holiday(group):
            weekday_hol  = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_holiday'] == 1)]
            saturday_hol = group[(group['day_str'].str.startswith('Sa', na=False))  & (group['is_holiday'] == 1)]
            return len(weekday_hol) * 1.0 + len(saturday_hol) * 0.5
        
        weighted_days    = subset_df.groupby('name').apply(calc_weighted_days_worked).reset_index()
        weighted_days.columns = ['name', 'weighted_days_worked']
        
        weighted_absence = subset_df.groupby('name').apply(calc_weighted_absence).reset_index()
        weighted_absence.columns = ['name', 'weighted_absence']

        weighted_leave   = subset_df.groupby('name').apply(calc_weighted_leave).reset_index()
        weighted_leave.columns = ['name', 'weighted_leave']

        weighted_holiday = subset_df.groupby('name').apply(calc_weighted_holiday).reset_index()
        weighted_holiday.columns = ['name', 'weighted_holiday']
        
        report = subset_df.groupby('name').agg({
            'is_day_worked': 'sum',
            'is_leave': 'sum',
            'is_holiday': 'sum',
            'is_unjustified_absence': 'sum',
            'daily_target_for_worked_day': 'sum', 
            'ENTRY > 10H': 'sum',
            'ENTRY > 8H': 'sum',
            'NO LUNCH': 'sum',
            'UNDER 9H': 'sum',
            'IS HALF DAY': 'sum',
            'hours_worked': 'sum',
            'daily_lunch_minutes': 'sum',
            'has_lunch_break': 'sum'
        }).reset_index()
        
        report = report.merge(weighted_days,    on='name', how='left')
        report = report.merge(weighted_absence, on='name', how='left')
        report = report.merge(weighted_leave,   on='name', how='left')
        report = report.merge(weighted_holiday, on='name', how='left')
        report = report.merge(starting_days,    on='name', how='left')
        
        report.rename(columns={
            'name': 'Employee name',
            'weighted_days_worked': 'days worked',
            'weighted_absence': 'ABSENCE',
            'daily_target_for_worked_day': 'TOTAL HOURS NEEDED',
            'hours_worked': 'TOTAL HOURS WORKED',
            'IS HALF DAY': 'HALF DAYS'
        }, inplace=True)

        saturday_records = subset_df[subset_df['day_str'].str.startswith('Sa', na=False)]
        expected_saturdays = len(saturday_records['day_numeric'].unique())

        # FIX: Use weighted leave/holiday so Saturday leaves count as 0.5
        report['real working days'] = (
            table_expected_days
            - report['weighted_leave']
            - report['weighted_holiday']
        )
        report.drop(['weighted_leave', 'weighted_holiday', 'is_leave', 'is_holiday'], axis=1, inplace=True)

        # Calculate actual worked hours (weekdays × 9h + saturdays × 5h)
        def calc_days_worked_hours(group):
            weekdays  = group[(~group['day_str'].str.startswith('Sa', na=False)) & (group['is_day_worked'] == 1)]
            saturdays = group[(group['day_str'].str.startswith('Sa', na=False))  & (group['is_day_worked'] == 1)]
            weekday_count  = len(weekdays['day_numeric'].unique())
            saturday_count = len(saturdays['day_numeric'].unique())
            return (weekday_count * WEEKDAY_HOURS) + (saturday_count * SATURDAY_HOURS)
        
        worked_hours_per_employee = subset_df.groupby('name').apply(calc_days_worked_hours).reset_index()
        worked_hours_per_employee.columns = ['name', 'hours worked']
        report = report.merge(worked_hours_per_employee, left_on='Employee name', right_on='name', how='left')
        if 'name' in report.columns:
            report.drop('name', axis=1, inplace=True)

        saturday_work_by_employee = saturday_records[saturday_records['is_day_worked'] == 1].groupby('name').size().reset_index(name='saturdays_worked')
        saturday_work_by_employee['saturdays_worked'] = saturday_work_by_employee['saturdays_worked'].apply(lambda x: 1 if x > 0 else 0)
        report = report.merge(saturday_work_by_employee, left_on='Employee name', right_on='name', how='left')
        report['saturdays_worked'] = report['saturdays_worked'].fillna(0)
        if 'name' in report.columns:
            report.drop('name', axis=1, inplace=True)
        report.drop(['saturdays_worked'], axis=1, inplace=True)
        
        report['avg_lunch_raw'] = report.apply(
            lambda x: x['daily_lunch_minutes'] / x['has_lunch_break'] if x['has_lunch_break'] > 0 else (x['daily_lunch_minutes'] if x['daily_lunch_minutes'] > 0 else 0), axis=1
        )
        report['AVG LUNCH TIME'] = report['avg_lunch_raw'].apply(minutes_to_hhmm)
        
        # Balance uses TOTAL HOURS WORKED (from Tps Eff, already accurate)
        report['balance_raw'] = report['TOTAL HOURS WORKED'] - report['TOTAL HOURS NEEDED']
        report['TOTAL HOURS NEEDED']      = report['TOTAL HOURS NEEDED'].apply(decimal_hours_to_hhmm)
        report['TOTAL HOURS WORKED']      = report['TOTAL HOURS WORKED'].apply(decimal_hours_to_hhmm)
        report['Balance of hours worked'] = report['balance_raw'].apply(decimal_hours_to_hhmm)
        
        report['STARTING DAY'] = report['starting_day'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else '')
        
        final_cols = [
            'Employee name',
            'STARTING DAY',
            'real working days',
            'days worked',
            'ABSENCE',
            'HALF DAYS',
            'UNDER 9H',
            'NO LUNCH',
            'AVG LUNCH TIME',
            'ENTRY > 10H',
            'ENTRY > 8H',
            'TOTAL HOURS NEEDED',
            'TOTAL HOURS WORKED',
            'Balance of hours worked'
        ]
        return report[final_cols].sort_values('ABSENCE', ascending=False)

    production_report = generate_production_report(df, working_days)

    # --- EXPORT ---
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            workbook  = writer.book
            worksheet = workbook.add_worksheet('Monthly Production Summary')
            
            header_title = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 14, 'font_color': '#2F5597', 'border': 1
            })
            section_header = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 12, 'fg_color': '#E7E6E6', 'border': 1
            })
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            header_red = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#C00000', 'font_color': 'white', 'border': 1
            })
            header_orange = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#ED7D31', 'font_color': 'white', 'border': 1
            })
            body_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
            text_format = workbook.add_format({'border': 1, 'align': 'left',   'valign': 'vcenter'})
            absence_red_format = workbook.add_format({
                'border': 1, 'align': 'center', 'valign': 'vcenter', 
                'bg_color': '#FFC7CE', 'font_color': '#9C0006'
            })
            
            num_cols    = 14
            current_row = 0
            
            worksheet.merge_range(0, 0, 0, num_cols - 1, header_text, header_title)
            current_row = 2
            
            if production_report is not None and not production_report.empty:
                table_start = min_date.strftime('%d/%m/%Y') if min_date else ''
                table_end   = max_date.strftime('%d/%m/%Y') if max_date else ''
                table_title = f"JOURS PRODUCTION (9h/jour, 5h Samedi) - Du {table_start} au {table_end}"
                
                worksheet.merge_range(current_row, 0, current_row, num_cols - 1, table_title, section_header)
                current_row += 1
                
                for col_num, value in enumerate(production_report.columns.values):
                    if "ABSENCE" in str(value):
                        worksheet.write(current_row, col_num, value, header_red)
                    elif "HALF DAYS" in str(value):
                        worksheet.write(current_row, col_num, value, header_orange)
                    else:
                        worksheet.write(current_row, col_num, value, header_format)
                current_row += 1
                
                for row_idx, row_data in production_report.iterrows():
                    for col_num, col_name in enumerate(production_report.columns):
                        value = row_data[col_name]
                        if pd.isna(value): 
                            value = ""
                        if col_name in ['real working days', 'days worked', 'ABSENCE', 'HALF DAYS', 'UNDER 9H', 'NO LUNCH', 'ENTRY > 10H', 'ENTRY > 8H']:
                            if value == 0: 
                                value = 0
                        elif col_name in ['AVG LUNCH TIME', 'Balance of hours worked', 'TOTAL HOURS WORKED']:
                            if value == 0 or value == "00:00": 
                                value = ""
                        if col_name == 'Employee name':
                            cell_fmt = text_format
                        else:
                            cell_fmt = body_format
                        if col_name == 'ABSENCE' and value > 0:
                            worksheet.write(current_row, col_num, value, absence_red_format)
                        else:
                            worksheet.write(current_row, col_num, value, cell_fmt)
                    current_row += 1
            
            for i in range(num_cols):
                if i == 0:
                    worksheet.set_column(i, i, 20)
                elif i == 1:
                    worksheet.set_column(i, i, 12)
                elif i in range(2, 11):
                    worksheet.set_column(i, i, 10)
                else:
                    worksheet.set_column(i, i, 14)

        print(f"\nSUCCESS! Monthly Production report generated: {output_path}")
        return output_path

    except Exception as e:
        print(f"Error saving file: {e}")
        import traceback
        print(f"DEBUG: {traceback.format_exc()}")
        return None

def main():
    if not os.path.exists(FOLDER_PATH):
        print("Folder not found.")
        return

    output = process_production_monthly_analysis(FOLDER_PATH, FOLDER_PATH)
    if output:
        print(f"Production Report generated: {output}")

if __name__ == "__main__":
    main()
