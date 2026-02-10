import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Suppress warnings from openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
EMPLOYEES_EXCLUS = [
    "HMOURI ALI"
]

CODES_OUVRIER = ['130', '140', '141', '131']

# --- HELPER CLASS FOR XLS COMPATIBILITY ---
class MockCell:
    """Mimics an openpyxl cell object for .xls files read via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Normalizes names to ensure matching works despite spaces/hidden chars."""
    if not name:
        return ""
    name = str(name).upper()
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

def extract_month_year_from_filename(file_path):
    """Extracts month and year from filename."""
    filename = os.path.basename(file_path).upper()
    
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    return '12', year

def extract_date_from_string(date_str):
    """Extract date from string like 'Lu 15/01/2025'"""
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(date_str))
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except Exception:
            return None
    return None

def get_sheet_rows(file_path):
    """Generator that yields rows from either .xlsx or .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Warning: '{os.path.basename(file_path)}' is an .xlsx file named as .xls. Switching engine...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Failed to read file with fallback: {e2}")
            else:
                print(f"Error processing .xls file {os.path.basename(file_path)}: {e}")
                return

def parse_scan_times(scan_str):
    """Parses scan string to extract all time entries."""
    if scan_str is None:
        return [], 0
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    return times, len(times)

def process_employee_buffer(employee_data):
    """Decides if an employee is an OUVRIER based on HJ codes."""
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    weekday_recs = [r for r in records if not str(r['day_str']).startswith(('Sa', 'Di'))]
    
    if not weekday_recs:
        return records

    ouvrier_matches = 0
    for r in weekday_recs:
        raw_hj = str(r.get('hj_code', ''))
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in CODES_OUVRIER:
            ouvrier_matches += 1
    
    ratio = ouvrier_matches / len(weekday_recs)
    if ratio > 0.5:
        return []
    
    return records

def extract_attendance_data_to_csv(file_path, output_csv_path):
    """
    Extracts attendance data from Excel file and saves it to a standardized CSV format.
    This avoids misinterpretation by having clean, structured data.
    """
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    source_file_name = os.path.basename(file_path)
    month_num, year_num = extract_month_year_from_filename(file_path)
    days_french = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''
            
            # Check for new section or employee (triggers buffer processing)
            if 'SERVICE / SECTION :' in val_0 or 'NOM :' in val_0:
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 'matricule': '', 'records': []
                }
            
            elif 'NOM :' in val_0:
                raw_name = val_0.replace('NOM :', '').strip()
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name),
                    'matricule': '', 
                    'records': []
                }

            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
                
            # Analyze daily data
            elif any(val_0.startswith(day) for day in days_french) and any(char.isdigit() for char in val_0):
                
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                
                # Check for leave/absence
                row_text = (val_0 + " " + str(raw_scan_val)).upper()
                if "CONGE-" in row_text:
                    continue

                if 'Date' not in val_0 and 'Heures' not in val_0:
                    times_list, scan_count = parse_scan_times(raw_scan_val)
                    
                    # Extract date information
                    date_obj = extract_date_from_string(val_0)
                    if date_obj:
                        day_numeric = date_obj.day
                        day_str = val_0.split()[0] if val_0 else ''
                        
                        # Determine day type and status
                        is_saturday = day_str.lower().startswith('sa')
                        is_sunday = day_str.lower().startswith('di')
                        
                        # Calculate basic metrics
                        hours_worked = 0.0
                        if times_list:
                            total_seconds = 0
                            for i in range(0, len(times_list) - 1, 2):
                                try:
                                    t_in = datetime.strptime(times_list[i], '%H:%M')
                                    t_out = datetime.strptime(times_list[i+1], '%H:%M')
                                    if t_out < t_in: t_out += timedelta(days=1)
                                    total_seconds += (t_out - t_in).total_seconds()
                                except:
                                    continue
                            hours_worked = round(total_seconds / 3600, 2)
                        
                        # Determine work status
                        is_leave = 0
                        is_holiday = 0
                        is_day_worked = 0
                        
                        if "JOUR FERIE" in row_text:
                            is_holiday = 1
                            if is_sunday: 
                                is_holiday = 0
                        elif "CONGE" in row_text:
                            is_leave = 1
                        elif hours_worked > 0:
                            is_day_worked = 1
                        
                        # Create standardized record
                        record = {
                            'source_file': source_file_name,
                            'service': current_employee.get('service', ''),
                            'name': current_employee.get('name', ''),
                            'matricule': current_employee.get('matricule', ''),
                            'full_date': date_obj.strftime('%Y-%m-%d') if date_obj else '',
                            'day_numeric': day_numeric,
                            'day_str': day_str,
                            'is_saturday': 1 if is_saturday else 0,
                            'is_sunday': 1 if is_sunday else 0,
                            'hj_code': str(hj_val).strip(),
                            'scan_count': scan_count,
                            'raw_pointages': str(raw_scan_val) if raw_scan_val else '',
                            'times_list': '|'.join(times_list),  # Store as pipe-separated string
                            'hours_worked': hours_worked,
                            'is_day_worked': is_day_worked,
                            'is_leave': is_leave,
                            'is_holiday': is_holiday,
                            'month_num': month_num,
                            'year_num': year_num
                        }
                        current_employee['records'].append(record)
        
        # Process final employee buffer
        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)

    except Exception as e:
        print(f"Error processing {os.path.basename(file_path)}: {e}")
        return False
    
    # Filter out excluded employees
    if EMPLOYEES_EXCLUS:
        excluded_clean = [clean_name_string(name) for name in EMPLOYEES_EXCLUS]
        all_records = [r for r in all_records if r['name'] not in excluded_clean]
    
    # Create DataFrame and save to CSV
    if all_records:
        df = pd.DataFrame(all_records)
        
        # Define column order for consistency
        column_order = [
            'source_file', 'service', 'name', 'matricule', 'full_date', 'day_numeric', 
            'day_str', 'is_saturday', 'is_sunday', 'hj_code', 'scan_count', 
            'raw_pointages', 'times_list', 'hours_worked', 'is_day_worked', 
            'is_leave', 'is_holiday', 'month_num', 'year_num'
        ]
        
        # Ensure all columns exist
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        
        df = df[column_order]
        
        # Save to CSV
        df.to_csv(output_csv_path, index=False, encoding='utf-8')
        print(f"✅ Successfully extracted {len(all_records)} records to {output_csv_path}")
        return True
    else:
        print(f"⚠️ No valid records found in {os.path.basename(file_path)}")
        return False

def process_all_excel_to_csv(input_dir, output_dir):
    """
    Processes all Excel files in input_dir and converts them to CSV format in output_dir.
    """
    if not os.path.exists(input_dir):
        print(f"Input directory not found: {input_dir}")
        return False
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    excel_files = [f for f in os.listdir(input_dir) 
                  if f.lower().endswith(('.xls', '.xlsx')) and not f.startswith("~$")]
    
    if not excel_files:
        print("No Excel files found in input directory")
        return False
    
    success_count = 0
    for excel_file in excel_files:
        input_path = os.path.join(input_dir, excel_file)
        csv_filename = os.path.splitext(excel_file)[0] + '.csv'
        output_path = os.path.join(output_dir, csv_filename)
        
        print(f"🔄 Processing {excel_file}...")
        if extract_attendance_data_to_csv(input_path, output_path):
            success_count += 1
    
    print(f"\n📊 Summary: {success_count}/{len(excel_files)} files successfully converted to CSV")
    return success_count > 0

if __name__ == "__main__":
    # Example usage
    input_directory = "temp_input"
    output_directory = "temp_csv"
    
    if process_all_excel_to_csv(input_directory, output_directory):
        print("CSV extraction completed successfully!")
    else:
        print("CSV extraction failed!")
