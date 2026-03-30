"""
pointage_pivot.py  (v2)
-----------------------
Reads one or more MEDIDIS "ETAT DES HEURES TRAVAILLEES" Excel files and
produces a styled pivot table showing per employee per month:
  - Date de début     : first ever pointage date for the employee
  - Heures travaillées: first entry → last exit each worked day
  - Heures requises   : working days × 9h (excl. Sundays unless worked,
                        excl. Moroccan public holidays unless worked)
  - Absences (jours)  : non-holiday weekdays where no timestamps exist

HJ codes skipped entirely: 505, 506, 131
Holidays treated like Sundays (not absent, not required) unless actually worked.

Usage:
    python pointage_pivot.py file1.xlsx [file2.xlsx ...] [-o output.xlsx]
"""

import re
import os
import argparse
from datetime import date, datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Morocco public holidays 2025
# ─────────────────────────────────────────────────────────────────────────────
MOROCCO_HOLIDAYS_2025 = {
    date(2025, 1,  1),   # Nouvel An
    date(2025, 1, 11),   # Manifeste de l'Independance
    date(2025, 1, 13),   # Aid Al Mawlid An Nabaoui (approx)
    date(2025, 3, 30),   # Aid Al Fitr J1
    date(2025, 3, 31),   # Aid Al Fitr J2
    date(2025, 4,  1),   # Aid Al Fitr J3
    date(2025, 5,  1),   # Fete du Travail
    date(2025, 6,  5),   # Aid Al Adha J1
    date(2025, 6,  6),   # Aid Al Adha J2
    date(2025, 6, 26),   # Aid Al Fitr 2025 (alternate, per govt)
    date(2025, 6, 27),   # Aid Al Adha 2025 (per govt calendar)
    date(2025, 7, 25),   # Fete du Trone
    date(2025, 7, 30),   # Allegiance des provinces du Sud
    date(2025, 8, 14),   # Recuperation de Oued Eddahab
    date(2025, 8, 20),   # Revolution du Roi et du Peuple
    date(2025, 8, 21),   # Fete de la Jeunesse
    date(2025, 9,  5),   # Aid Al Mawlid An Nabaoui 2025
    date(2025,11,  6),   # Marche Verte
    date(2025,11, 18),   # Fete de l'Independance
}

SKIP_HJ = {'505', '506', '131'}
REQUIRED_HOURS_PER_DAY = 9.0

MONTHS_FR = [
    "Janvier","Fevrier","Mars","Avril","Mai","Juin",
    "Juillet","Aout","Septembre","Octobre","Novembre","Decembre"
]

C_NAVY   = "1F3864"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_ORANGE = "FCE4D6"
C_ALTROW = "EBF3FB"
C_WHITE  = "FFFFFF"
C_GREEN  = "D9EAD3"
C_RED    = "FCE4D6"
C_YELLOW = "FFF2CC"

TIME_RE = re.compile(r'\b(\d{2}:\d{2})\b')
DAY_RE  = re.compile(r'^(Lu|Ma|Me|Je|Ve|Sa|Di)\s+(\d{2}/\d{2}/\d{4})$')


def parse_first_last(pointage):
    if not pointage or not isinstance(pointage, str):
        return None, None
    times = TIME_RE.findall(pointage)
    return (times[0], times[-1]) if times else (None, None)


def to_min(t):
    h, m = map(int, t.split(':'))
    return h * 60 + m


def calc_hours(first, last):
    if not first or not last:
        return 0.0
    f, l = to_min(first), to_min(last)
    diff = (1440 - f + l) if l < f else (l - f)
    return round(diff / 60, 4)


def is_day_off(d: date) -> bool:
    return d.weekday() == 6 or d in MOROCCO_HOLIDAYS_2025


def extract_employees(rows):
    employees = []
    current_emp = None
    current_service = ""
    in_data = False

    for row in rows:
        c0 = row[0] if len(row) > 0 else None
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None

        if not c0 and not c1:
            in_data = False
            continue

        if isinstance(c0, str) and c0.startswith("SERVICE / SECTION :"):
            current_service = c0.replace("SERVICE / SECTION :", "").strip()
            in_data = False
            continue

        if isinstance(c0, str) and c0.startswith("NOM :"):
            if current_emp is not None:
                employees.append(current_emp)
            current_emp = {
                "name": c0.replace("NOM :", "").strip(),
                "matricule": "",
                "service": current_service,
                "days": [],
                "first_date": None,
            }
            in_data = False
            continue

        if isinstance(c0, str) and c0.startswith("MATRICULE :") and current_emp:
            current_emp["matricule"] = c0.replace("MATRICULE :", "").strip()
            continue

        if isinstance(c0, str) and c0 == "Date":
            in_data = True
            continue

        if in_data and current_emp and isinstance(c0, str):
            m = DAY_RE.match(c0.strip())
            if m:
                date_str = m.group(2)
                try:
                    dt = datetime.strptime(date_str, "%d/%m/%Y").date()
                except ValueError:
                    continue

                pointage = c2
                first, last = parse_first_last(pointage)
                hours = calc_hours(first, last)
                # Check if pointage contains "Absence" or has no timestamps
                is_absence_marked = isinstance(pointage, str) and "absence" in pointage.lower()
                has_work = bool(first) and not is_absence_marked

                # Track first date from ANY valid timestamp (before HJ filtering)
                # to get the actual first badge scan overall
                if first and (current_emp["first_date"] is None or dt < current_emp["first_date"]):
                    current_emp["first_date"] = dt

                # Skip HJ codes for calculations, but first_date is already tracked above
                hj_code = str(c1).strip() if c1 else ""
                if hj_code in SKIP_HJ:
                    continue

                current_emp["days"].append({
                    "date": dt,
                    "month": dt.month,
                    "hours": hours,
                    "has_work": has_work,
                    "is_day_off": is_day_off(dt),
                })
            else:
                in_data = False

    if current_emp is not None:
        employees.append(current_emp)

    return employees


def aggregate_monthly(employees):
    result = []
    for emp in employees:
        monthly = {mo: {"hours": 0.0, "required_hours": 0.0, "absent_days": 0} for mo in range(1, 13)}

        for day in emp["days"]:
            mo = day["month"]
            if day["has_work"]:
                monthly[mo]["hours"] += day["hours"]
                monthly[mo]["required_hours"] += REQUIRED_HOURS_PER_DAY
            else:
                if not day["is_day_off"]:
                    monthly[mo]["absent_days"] += 1
                    monthly[mo]["required_hours"] += REQUIRED_HOURS_PER_DAY

        result.append({
            "name": emp["name"],
            "matricule": emp["matricule"],
            "service": emp["service"],
            "first_date": emp["first_date"],
            "monthly": monthly,
        })
    return result


def hfill(c): return PatternFill("solid", fgColor=c)

def bdr():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

def lft(): return Alignment(horizontal="left", vertical="center")

def wcell(ws, r, c, v, font=None, fill=None, align=None, border=None, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if nf:     cell.number_format = nf
    return cell


def write_pivot_sheet(ws, data, sheet_title, year=None):
    FIXED = 3
    MONTH_COLS = 3
    TOTAL_COLS = 3
    TOTAL_WIDTH = FIXED + 12 * MONTH_COLS + TOTAL_COLS

    # Extract year from data if not provided
    if year is None and data:
        for emp in data:
            if emp.get("first_date"):
                year = emp["first_date"].year
                break
            # Check monthly data for year
            for mo in range(1, 13):
                if emp.get("monthly", {}).get(mo, {}).get("hours", 0) > 0:
                    # Try to find year from any day data if available
                    pass
    if year is None:
        year = datetime.now().year

    wf  = Font(color=C_WHITE, bold=True, size=10)
    df  = Font(color="000000", size=9)
    dbf = Font(color="000000", bold=True, size=9)
    b   = bdr()

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_WIDTH)
    wcell(ws, 1, 1, f"MEDIDIS — {sheet_title} — Pointage {year}",
          font=Font(bold=True, size=13, color=C_WHITE),
          fill=hfill(C_NAVY), align=ctr())

    # Row 3: group headers
    wcell(ws, 3, 1, "Matricule", font=wf, fill=hfill(C_NAVY), align=ctr(), border=b)
    wcell(ws, 3, 2, "Nom",       font=wf, fill=hfill(C_NAVY), align=ctr(), border=b)
    wcell(ws, 3, 3, "Debut",     font=wf, fill=hfill(C_NAVY), align=ctr(), border=b)

    for mo in range(1, 13):
        cs = FIXED + (mo - 1) * MONTH_COLS + 1
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs + MONTH_COLS - 1)
        wcell(ws, 3, cs, MONTHS_FR[mo - 1],
              font=wf, fill=hfill(C_BLUE), align=ctr(), border=b)

    tc = FIXED + 12 * MONTH_COLS + 1
    ws.merge_cells(start_row=3, start_column=tc, end_row=3, end_column=tc + TOTAL_COLS - 1)
    wcell(ws, 3, tc, "TOTAL ANNUEL", font=wf, fill=hfill(C_NAVY), align=ctr(), border=b)

    # Row 4: sub-headers
    for fc in range(1, 4):
        wcell(ws, 4, fc, "", fill=hfill(C_NAVY), border=b)

    for mo in range(1, 13):
        cs = FIXED + (mo - 1) * MONTH_COLS + 1
        wcell(ws, 4, cs,     "H.Trav.", font=Font(bold=True, size=8, color="1A4D00"), fill=hfill(C_LBLUE), align=ctr(), border=b)
        wcell(ws, 4, cs + 1, "H.Req.",  font=Font(bold=True, size=8, color="1F3864"), fill=hfill(C_LBLUE), align=ctr(), border=b)
        wcell(ws, 4, cs + 2, "Abs(j)",  font=Font(bold=True, size=8, color="7F0000"), fill=hfill(C_LBLUE), align=ctr(), border=b)

    wcell(ws, 4, tc,     "H.Trav.", font=Font(bold=True, size=9, color="1A4D00"), fill=hfill(C_ORANGE), align=ctr(), border=b)
    wcell(ws, 4, tc + 1, "H.Req.",  font=Font(bold=True, size=9, color="1F3864"), fill=hfill(C_ORANGE), align=ctr(), border=b)
    wcell(ws, 4, tc + 2, "Abs(j)",  font=Font(bold=True, size=9, color="7F0000"), fill=hfill(C_ORANGE), align=ctr(), border=b)

    ws.freeze_panes = ws.cell(row=5, column=4)

    # Data rows
    for idx, emp in enumerate(data):
        row = 5 + idx
        rfill = hfill(C_ALTROW) if idx % 2 == 0 else hfill(C_WHITE)
        first_date_val = emp["first_date"].strftime("%d/%m/%Y") if emp["first_date"] else "—"

        wcell(ws, row, 1, emp["matricule"],  font=df,  fill=rfill, align=ctr(), border=b)
        wcell(ws, row, 2, emp["name"],       font=dbf, fill=rfill, align=lft(), border=b)
        wcell(ws, row, 3, first_date_val,    font=df,  fill=rfill, align=ctr(), border=b)

        tot_h = tot_r = tot_a = 0.0

        for mo in range(1, 13):
            cs  = FIXED + (mo - 1) * MONTH_COLS + 1
            h   = emp["monthly"][mo]["hours"]
            req = emp["monthly"][mo]["required_hours"]
            ab  = emp["monthly"][mo]["absent_days"]
            tot_h += h; tot_r += req; tot_a += ab

            wcell(ws, row, cs,     round(h,   2), font=df, fill=hfill(C_GREEN),  align=ctr(), border=b, nf='0.00')
            wcell(ws, row, cs + 1, round(req, 2), font=df, fill=hfill(C_YELLOW), align=ctr(), border=b, nf='0.00')
            wcell(ws, row, cs + 2, int(ab),       font=df, fill=hfill(C_RED),    align=ctr(), border=b)

        wcell(ws, row, tc,     round(tot_h, 2), font=dbf, fill=hfill(C_ORANGE), align=ctr(), border=b, nf='0.00')
        wcell(ws, row, tc + 1, round(tot_r, 2), font=dbf, fill=hfill(C_ORANGE), align=ctr(), border=b, nf='0.00')
        wcell(ws, row, tc + 2, int(tot_a),      font=dbf, fill=hfill(C_ORANGE), align=ctr(), border=b)

    # Column widths
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    for mo in range(1, 13):
        cs = FIXED + (mo - 1) * MONTH_COLS + 1
        ws.column_dimensions[get_column_letter(cs)].width     = 8
        ws.column_dimensions[get_column_letter(cs + 1)].width = 8
        ws.column_dimensions[get_column_letter(cs + 2)].width = 7
    for i in range(3):
        ws.column_dimensions[get_column_letter(tc + i)].width = 10

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20
    for i in range(len(data)):
        ws.row_dimensions[5 + i].height = 16

    ws.sheet_view.showGridLines = False


def process_file(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return aggregate_monthly(extract_employees(rows))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("inputs", nargs="*",
                        default=["SEC_BIR2_2025.xlsx", "SECTBIR_1__PRODUCTION2025_xls.xlsx"])
    parser.add_argument("-o", "--output", default="pointage_pivot_output.xlsx")
    args = parser.parse_args()

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    all_data = []

    for path in args.inputs:
        print(f"Processing: {path}")
        data = process_file(path)
        print(f"  -> {len(data)} employees found")
        all_data.extend(data)
        base = os.path.splitext(os.path.basename(path))[0][:28]
        ws = out_wb.create_sheet(title=base)
        write_pivot_sheet(ws, data, sheet_title=base)

    if len(args.inputs) > 1:
        ws_all = out_wb.create_sheet(title="TOUS")
        write_pivot_sheet(ws_all, all_data, sheet_title="Tous les employes")
        print(f"  -> Merged sheet: {len(all_data)} total employees")

    out_wb.save(args.output)
    print(f"\nOutput saved to: {args.output}")


def process_annual_pivot(input_dir, output_dir):
    """
    Process annual pivot analysis for Streamlit integration.
    Returns the output file path or None if no files processed.
    """
    # Get all Excel files from input directory
    input_files = []
    for f in os.listdir(input_dir):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            input_files.append(os.path.join(input_dir, f))
    
    print(f"[DEBUG] Found {len(input_files)} files in input directory")
    
    if not input_files:
        return None
    
    # Generate output filename
    output_path = os.path.join(output_dir, "Annual_Pivot_Report.xlsx")
    
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    all_data = []

    for path in input_files:
        filename = os.path.basename(path)
        print(f"[DEBUG] Processing: {filename}")
        
        data = process_file(path)
        print(f"[DEBUG] Employees found: {len(data) if data else 0}")
        
        if not data:
            continue
        
        all_data.extend(data)
        base = os.path.splitext(os.path.basename(path))[0][:28]
        ws = out_wb.create_sheet(title=base)
        write_pivot_sheet(ws, data, sheet_title=base)

    if len(input_files) > 1 and all_data:
        ws_all = out_wb.create_sheet(title="TOUS")
        write_pivot_sheet(ws_all, all_data, sheet_title="Tous les employes")
        print(f"[DEBUG] Merged sheet: {len(all_data)} total employees")

    out_wb.save(output_path)
    print(f"[DEBUG] Saved: {output_path}")
    
    return output_path


if __name__ == "__main__":
    main()
