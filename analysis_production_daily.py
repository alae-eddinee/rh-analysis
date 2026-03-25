import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Supprimer les avertissements de openpyxl si il lit des fichiers mal nommés
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
CHEMIN_DOSSIER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")
NOM_FICHIER_SORTIE = "Analyse_Quotidienne_Production.xlsx"

# LISTE DES EMPLOYÉS À EXCLURE PAR NOM (Insensible à la casse)
EMPLOYES_EXCLUS = [
    # "ABOU HASNAA", 
    "HMOURI ALI"
]

# PRODUCTION WORKER CODES - Only process employees with these codes
PRODUCTION_CODES = ['130', '131', '140', '141']

# Hours per day for production workers
WEEKDAY_HOURS = 9.0
SATURDAY_HOURS = 5.0

# --- CLASSE UTILITAIRE POUR COMPATIBILITÉ XLS ---
class MockCell:
    """Imite un objet cellule openpyxl pour les fichiers .xls lus via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Minimal cleaning - preserve original case and spacing."""
    if not name:
        return ""
    name = str(name)
    # Only remove problematic whitespace characters but preserve original format
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    # Don't uppercase, don't normalize multiple spaces - keep original name
    return name.strip()

def parse_scan_times(scan_str):
    """Analyse la chaîne pour trouver toutes les entrées de temps (HH:MM)."""
    if scan_str is None:
        return {}, 0, []
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    count = len(times)
    scans = {}
    for i, time_val in enumerate(times):
        scans[f'scan_{i+1}'] = time_val  
    return scans, count, times

def get_sheet_rows(file_path):
    """Générateur qui produit des lignes de fichiers .xlsx ou .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Attention : '{os.path.basename(file_path)}' est un fichier .xlsx nommé comme .xls. Changement de moteur...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Échec de lecture du fichier avec secours : {e2}")
            else:
                print(f"Erreur lors du traitement du fichier .xls {os.path.basename(file_path)} : {e}")
                return

def process_employee_buffer(employee_data):
    """Décide si un employé est un Production worker basé sur les codes HJ."""
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    name = employee_data.get('name', 'Unknown')
    
    weekday_recs = []
    for r in records:
        day_str = str(r.get('day_str', '')).lower()
        if not day_str.startswith('sa') and not day_str.startswith('di'):
            weekday_recs.append(r)
    
    if not weekday_recs:
        return records

    prod_matches = 0
    for r in weekday_recs:
        raw_hj = str(r['hj_code'])
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in PRODUCTION_CODES:
            prod_matches += 1
    
    ratio = prod_matches / len(weekday_recs)
    # Production analysis: ONLY include if more than 50% of codes are production codes
    if ratio > 0.5:
        return records
    
    return []

def extract_month_year_from_filename(file_path):
    """Extrait le mois et l'année du nom de fichier."""
    filename = os.path.basename(file_path).upper()
    
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    return '12', year

def is_friday(date_obj):
    """Check if date is Friday."""
    if date_obj is None:
        return False
    return date_obj.weekday() == 4

def extract_date_from_string(date_str):
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(date_str))
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except:
            return None
    return None

def extract_daily_data(file_path):
    """Extrait les données, met en mémoire tampon par employé pour vérifier le statut TAP via la colonne HJ."""
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    source_file_name = os.path.basename(file_path)
    month_num, year_num = extract_month_year_from_filename(file_path)
    days_french = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''
            
            if 'SERVICE / SECTION :' in val_0 or val_0.upper().startswith('NOM :'):
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 'matricule': '', 'records': []
                }
            
            elif val_0.upper().startswith('NOM :'):
                raw_name = val_0.split(':', 1)[1].strip() if ':' in val_0 else val_0
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name),
                    'matricule': '',
                    'records': []
                }

            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
                
            elif any(val_0.startswith(day) for day in days_french) and any(char.isdigit() for char in val_0):
                
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                
                row_text = (val_0 + " " + str(raw_scan_val)).upper()
                if "CONGE-" in row_text:
                    continue

                if 'Date' not in val_0 and 'Heures' not in val_0:
                    scan_times_dict, calculated_count, times_list = parse_scan_times(raw_scan_val)
                    parts = val_0.split()
                    
                    day_match = re.search(r'\d+', val_0)
                    day_num = int(day_match.group()) if day_match else 0
                    day_str = parts[0] if parts else ''
                    
                    full_date = extract_date_from_string(val_0)
                    if full_date is None:
                        try:
                            full_date = datetime(int(year_num), int(month_num), day_num)
                        except:
                            full_date = None

                    record = {
                        'source_file': source_file_name,
                        'name': current_employee.get('name', ''),
                        'day_raw': parts[0] if parts else '',
                        'day_numeric': day_num,
                        'day_str': day_str,
                        'full_date': full_date,
                        'hj_code': str(hj_val).strip(),
                        'scan_count': calculated_count,
                        'raw_pointages': str(raw_scan_val) if raw_scan_val else '',
                        'month_num': month_num,
                        'year_num': year_num
                    }
                    current_employee['records'].append(record)
        
        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)

    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier {os.path.basename(file_path)} : {e}")
        return []
    
    return all_records

def analyze_row(row):
    """Calcule les indicateurs pour retard, pas de déjeuner, heures et demi-journée pour TAP."""
    scans = re.findall(r'\d{1,2}:\d{2}', str(row.get('raw_pointages', '')))
    
    late_800 = False
    late_1000 = False
    no_lunch = False
    is_half_day = False
    hours_worked = 0.0
    is_absent = False

    if not scans:
        is_absent = True
        return late_800, late_1000, no_lunch, hours_worked, is_half_day, is_absent
    
    total_seconds = 0
    for i in range(0, len(scans) - 1, 2):
        t_in = datetime.strptime(scans[i], '%H:%M')
        t_out = datetime.strptime(scans[i+1], '%H:%M')
        if t_out < t_in: t_out += timedelta(days=1)
        total_seconds += (t_out - t_in).total_seconds()
        
    hours_worked = round(total_seconds / 3600, 2)
    
    first_scan_dt = datetime.strptime(scans[0], '%H:%M')
    
    limit_800 = first_scan_dt.replace(hour=8, minute=0, second=0)
    limit_1000 = first_scan_dt.replace(hour=10, minute=0, second=0)
    limit_1300 = first_scan_dt.replace(hour=13, minute=0, second=0)

    if first_scan_dt > limit_1000:
        late_1000 = True
        late_800 = False
    elif first_scan_dt > limit_800:
        late_1000 = False
        late_800 = True

    day_str = str(row.get('day_str', '')).lower()
    is_saturday = day_str.startswith('sa')
    is_friday_flag = is_friday(row.get('full_date'))

    # TAP: Check lunch break (4 scans needed)
    if is_saturday:
        no_lunch = False
    else:
        no_lunch = len(scans) < 4 and len(scans) > 0

    if not is_saturday and len(scans) >= 2 and hours_worked > 0:
        t_first = datetime.strptime(scans[0], '%H:%M')
        t_last = datetime.strptime(scans[-1], '%H:%M')
        
        if t_last < t_first: 
            t_last += timedelta(days=1)
            
        cond_afternoon = (t_first >= limit_1300)
        cond_morning = (t_last <= limit_1300) and (hours_worked < 8.0)
        
        if cond_afternoon or cond_morning:
            is_half_day = True
    
    return late_800, late_1000, no_lunch, hours_worked, is_half_day, is_absent

def create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, flag_column, output_header):
    """Crée un DataFrame à 3 colonnes : [Nom, Compte, %]"""
    subset = daily_df[daily_df[flag_column]].copy()
    result = subset[['name']].reset_index(drop=True)
    
    if flag_column == 'is_under_hours' and daily_df['day_str'].iloc[0].startswith('Sa'):
        stats_to_use = monthly_stats_saturday
    else:
        stats_to_use = monthly_stats
    
    result['Count'] = result['name'].map(stats_to_use[flag_column]).fillna(0).astype(int)
    
    total_days = result['name'].map(stats_to_use['total_attendance']).fillna(1) 
    
    result['%'] = (result['Count'] / total_days)
    
    result.columns = [output_header, 'Count', '%']
    return result

def process_production_daily_analysis(input_dir, output_dir):
    """Traite les fichiers TAP dans input_dir et sauvegarde l'analyse dans output_dir."""
    if not os.path.exists(input_dir):
        print(f"Dossier non trouvé : {input_dir}")
        return None
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []

    print("Analyse des fichiers Production...")
    for file in os.listdir(input_dir):
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Lecture : {file}...")
            full_path = os.path.join(input_dir, file)
            records = extract_daily_data(full_path)
            all_data.extend(records)

    if not all_data:
        print("DEBUG: No data extracted - check PRODUCTION_CODES filter")
        print(f"PRODUCTION_CODES = {PRODUCTION_CODES}")
        return None

    df = pd.DataFrame(all_data)

    if EMPLOYES_EXCLUS:
        print(f"\nFiltrage des noms exclus : {EMPLOYES_EXCLUS}")
        excluded_clean = [clean_name_string(name) for name in EMPLOYES_EXCLUS]
        initial_len = len(df)
        df = df[~df['name'].isin(excluded_clean)]
        print(f"Supprimé {initial_len - len(df)} enregistrements basés sur la liste d'exclusion de noms.")
    
    if df.empty:
        print(f"DEBUG: DataFrame empty after filtering. Initial records: {len(all_data)}")
        return None

    # --- DÉTECTION CHRONOLOGIQUE ---
    if 'day_numeric' in df.columns and not df.empty:
        month_num = df['month_num'].iloc[0] if 'month_num' in df.columns else '01'
        year_num = df['year_num'].iloc[0] if 'year_num' in df.columns else '2026'
        
        unique_days_in_order = []
        seen = set()
        for d in df['day_numeric']:
            if d not in seen:
                unique_days_in_order.append(d)
                seen.add(d)

        real_start_day = unique_days_in_order[0]
        real_end_day = unique_days_in_order[-1]
        
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days_in_order) - 1):
            if unique_days_in_order[i] > unique_days_in_order[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        print(f"\n--- ANALYSE DE LA PÉRIODE ---")
        print(f"Séquence détectée : {unique_days_in_order}")
        
        target_report_day = real_end_day
        
        last_day_records = df[df['day_numeric'] == target_report_day]
        total_last_day = len(last_day_records)
        incomplete_count = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total_last_day > 0 and (incomplete_count / total_last_day) > 0.5:
            print(f"DÉCISION : Le jour {target_report_day} est incomplet (en cours).")
            df = df[df['day_numeric'] != target_report_day].copy()
            if len(unique_days_in_order) > 1:
                target_report_day = unique_days_in_order[-2]
                real_end_day = target_report_day
            print(f"Nouveau jour cible : {target_report_day}")
        else:
            print(f"DÉCISION : Le jour {target_report_day} est complet.")

        month_names = {
            '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
            '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
            '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
        }
        month_name = month_names.get(month_num, f'Mois {month_num}')
        
    else:
        print("DEBUG: No day_numeric column found in data")
        print(f"DEBUG: Available columns: {df.columns.tolist()}")
        return None

    print("\nCalcul des métriques Production...")
    results = df.apply(analyze_row, axis=1)
    
    df['is_late_800'] = [x[0] for x in results]
    df['is_late_1000'] = [x[1] for x in results]
    df['no_lunch'] = [x[2] for x in results]
    df['hours_worked'] = [x[3] for x in results]
    df['is_half_day'] = [x[4] for x in results] 
    df['is_absent'] = [x[5] for x in results] 
    
    if 'day_str' in df.columns:
        mask_saturday = df['day_str'].astype(str).str.startswith('Sa')
        df.loc[mask_saturday, 'no_lunch'] = False

    # TAP: 9h target for weekdays, 5h for Saturdays
    df['target_hours'] = df.apply(lambda x: 5.0 if str(x['day_str']).startswith('Sa') else 9.0, axis=1)
    df['is_under_hours'] = (df['scan_count'] > 0) & (df['hours_worked'] < df['target_hours'])

    cols_to_sum = ['is_late_800', 'is_late_1000', 'no_lunch', 'is_half_day', 'is_absent']
    
    valid_days_df = df[df['hours_worked'] > 0]
    
    saturday_records = valid_days_df[valid_days_df['day_str'].str.startswith('Sa')]
    weekday_records = valid_days_df[~valid_days_df['day_str'].str.startswith('Sa')]
    
    monthly_stats_weekday = weekday_records.groupby('name')[cols_to_sum].sum()
    monthly_stats_weekday['total_attendance'] = weekday_records.groupby('name').size()
    monthly_stats_weekday['is_under_hours'] = weekday_records.groupby('name')['is_under_hours'].sum()
    
    monthly_stats_saturday = saturday_records.groupby('name')[cols_to_sum].sum()
    monthly_stats_saturday['total_attendance'] = saturday_records.groupby('name').size()
    monthly_stats_saturday['is_under_hours'] = saturday_records.groupby('name')['is_under_hours'].sum()
    
    monthly_stats = monthly_stats_weekday.combine_first(monthly_stats_saturday)
    
    # --- GET TARGET DAY FOR DAILY ANALYSIS ---
    # Get actual available days from the (possibly filtered) df
    available_days = df['day_numeric'].unique()
    print(f"DEBUG: Available days after filtering: {available_days}")
    if len(available_days) == 0:
        print("DEBUG: No days available after filtering")
        return None
    
    # Use the last available day in the filtered df
    target_report_day = max(available_days)
    
    if 'day_numeric' in df.columns:
        daily_df = df[df['day_numeric'] == target_report_day].copy()
    else:
        daily_df = pd.DataFrame()

    if daily_df.empty:
        print(f"DEBUG: daily_df empty for target day {target_report_day}")
        print(f"DEBUG: df has {len(df)} records, day_numeric values: {df['day_numeric'].unique()}")
        return None

    sample_day_str = daily_df.iloc[0]['day_str'] if not daily_df.empty else ""
    is_target_saturday = str(sample_day_str).startswith('Sa')

    if is_target_saturday:
        under_header = "Moins de 5h"
    else:
        under_header = "Moins de 9h"
    
    df_under = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_under_hours', under_header)
    df_late_10 = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_late_1000', "Entrée > 10:00")
    df_late_8 = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_late_800', "Entrée > 08:00")
    df_half_day = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_half_day', "Demi-Journée")
    df_absent = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_absent', "Absence")

    if is_target_saturday:
        main_list = pd.concat([df_under, df_late_10, df_late_8, df_absent], axis=1)
    else:
        df_no_lunch = create_category_dataframe(daily_df[~daily_df['is_half_day']], monthly_stats, monthly_stats_saturday, 'no_lunch', "Pas de Déjeuner")
        main_list = pd.concat([df_under, df_half_day, df_no_lunch, df_late_10, df_late_8, df_absent], axis=1)

    if not df.empty and 'day_numeric' in df.columns:
        if has_transition:
            first_month_days = unique_days_in_order[:pivot_index + 1]
            second_month_days = unique_days_in_order[pivot_index + 1:]
            total_days = len(first_month_days) + len(second_month_days)
        else:
            total_days = len(unique_days_in_order)
        
        dynamic_filename = f"POINTAGE PRODUCTION ANALYSE DU {real_start_day:02d}-{month_num}-{year_num} A {real_end_day:02d}-{month_num}-{year_num}.xlsx"
        output_path = os.path.join(output_dir, dynamic_filename)
        
        header_text = f"Analyse Quotidienne Production - Période : {real_start_day} au {real_end_day} {month_name} {year_num}"
    else:
        header_text = "Analyse Quotidienne Production - Période non spécifiée"
        output_path = os.path.join(output_dir, NOM_FICHIER_SORTIE)
    
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            main_list.to_excel(writer, sheet_name='Analyse Production Quotidienne', index=False, header=False, startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['Analyse Production Quotidienne']
            
            header_title = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 14, 'font_color': '#2F5597', 'border': 1
            })
            
            if len(main_list.columns) > 1:
                worksheet.merge_range(0, 0, 0, len(main_list.columns) - 1, header_text, header_title)
            else:
                worksheet.write(0, 0, header_text, header_title)
            
            header_blue = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            header_orange = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#ED7D31', 'font_color': 'white', 'border': 1
            })
            header_red = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#C00000', 'font_color': 'white', 'border': 1
            })
            
            body_left = workbook.add_format({'border': 1, 'align': 'left'})
            body_center = workbook.add_format({'border': 1, 'align': 'center'})
            body_pct = workbook.add_format({'border': 1, 'align': 'center', 'num_format': '0%'})

            max_rows = len(main_list)
            columns = main_list.columns.tolist()

            for i, col_name in enumerate(columns):
                col_name_str = str(col_name)
                
                col_format = body_left
                header_style = header_blue

                if "Count" in col_name_str:
                    header_style = header_orange
                    col_format = body_center
                elif "%" in col_name_str:
                    header_style = header_orange
                    col_format = body_pct
                elif "Demi-Journée" in col_name_str:
                    header_style = header_orange
                elif "Absence" in col_name_str:
                    header_style = header_red
                
                worksheet.write(1, i, col_name, header_style)

                col_data = main_list.iloc[:, i]
                max_data_len = 0
                if "%" in col_name_str:
                    max_data_len = 5
                else:
                    valid_data = col_data.dropna().astype(str)
                    if not valid_data.empty:
                        max_data_len = valid_data.map(len).max()
                
                final_width = max(max_data_len, len(col_name_str)) + 4
                worksheet.set_column(i, i, final_width)

                for row_idx in range(max_rows):
                    cell_val = main_list.iloc[row_idx, i]
                    if pd.isna(cell_val):
                        worksheet.write(row_idx + 2, i, "", col_format)
                    else:
                        worksheet.write(row_idx + 2, i, cell_val, col_format)

        print(f"\nSUCCÈS ! Rapport Production sauvegardé : {output_path}")
        return output_path

    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier : {e}")
        return None

def main():
    if not os.path.exists(CHEMIN_DOSSIER):
        print(f"Dossier non trouvé : {CHEMIN_DOSSIER}")
        return
    
    output = process_production_daily_analysis(CHEMIN_DOSSIER, CHEMIN_DOSSIER)
    if output:
        print(f"Fichier généré : {output}")

if __name__ == "__main__":
    main()
